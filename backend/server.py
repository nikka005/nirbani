from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import HTMLResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from bson import ObjectId

# Import services
from sms_service import send_collection_sms, send_payment_sms
from bill_service import generate_farmer_bill_html, generate_daily_report_html

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'nirbani-dairy-secret-key-2026')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="Nirbani Dairy Management System")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

# User Models
class UserCreate(BaseModel):
    name: str
    email: EmailStr
    phone: str
    password: str
    role: str = "staff"  # admin, staff, branch_manager

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    email: str
    phone: str
    role: str
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

# Farmer Models
class FarmerCreate(BaseModel):
    name: str
    phone: str
    address: Optional[str] = ""
    village: Optional[str] = ""
    bank_account: Optional[str] = ""
    ifsc_code: Optional[str] = ""
    aadhar_number: Optional[str] = ""
    milk_type: Optional[str] = "cow"  # cow, buffalo, both
    fixed_rate: Optional[float] = None  # legacy single rate
    cow_rate: Optional[float] = None
    buffalo_rate: Optional[float] = None

class FarmerResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    address: str
    village: str
    bank_account: str
    ifsc_code: str
    aadhar_number: str
    milk_type: str = "cow"
    fixed_rate: Optional[float] = None
    cow_rate: Optional[float] = None
    buffalo_rate: Optional[float] = None
    total_milk: float
    total_due: float
    total_paid: float
    balance: float
    created_at: str
    is_active: bool

class FarmerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    village: Optional[str] = None
    bank_account: Optional[str] = None
    ifsc_code: Optional[str] = None
    aadhar_number: Optional[str] = None
    is_active: Optional[bool] = None
    milk_type: Optional[str] = None
    fixed_rate: Optional[float] = None
    cow_rate: Optional[float] = None
    buffalo_rate: Optional[float] = None

# Milk Collection Models
class MilkCollectionCreate(BaseModel):
    farmer_id: str
    shift: str  # morning, evening
    quantity: float  # in liters
    fat: float  # fat percentage
    snf: Optional[float] = None  # SNF percentage
    milk_type: Optional[str] = None  # cow, buffalo, mix (auto from farmer if not set)
    date: Optional[str] = None  # optional date override (YYYY-MM-DD) for historical entries
    rate: Optional[float] = None  # optional rate override

class MilkCollectionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    farmer_id: str
    farmer_name: str
    shift: str
    quantity: float
    fat: float
    snf: float
    rate: float
    amount: float
    milk_type: str = "cow"
    date: str
    created_at: str

# Rate Chart Models
class RateChartEntry(BaseModel):
    fat: float
    snf: float
    rate: float

class RateChartCreate(BaseModel):
    name: str
    entries: List[RateChartEntry]
    is_default: bool = False

class RateChartResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    entries: List[dict]
    is_default: bool
    created_at: str
    updated_at: str

# Payment Models
class PaymentCreate(BaseModel):
    farmer_id: str
    amount: float
    payment_mode: str  # cash, upi, bank
    payment_type: str = "payment"  # payment, advance, deduction
    notes: Optional[str] = ""

class PaymentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    farmer_id: str
    farmer_name: str
    amount: float
    payment_mode: str
    payment_type: str = "payment"
    notes: str = ""
    date: str
    created_at: str

# Customer Models
class CustomerCreate(BaseModel):
    name: str
    phone: str
    address: Optional[str] = ""
    customer_type: str = "retail"  # retail, wholesale
    gst_number: Optional[str] = ""

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    customer_type: Optional[str] = None
    gst_number: Optional[str] = None


class CustomerResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    address: str
    customer_type: str
    gst_number: str
    total_purchase: float
    total_paid: float
    balance: float
    created_at: str
    is_active: bool

# Sale Models
class SaleCreate(BaseModel):
    customer_id: str
    product: str  # milk, paneer, dahi, ghee, lassi
    quantity: Optional[float] = 0
    rate: Optional[float] = 0
    direct_amount: Optional[float] = None
    notes: Optional[str] = ""
    date: Optional[str] = None  # optional date override (YYYY-MM-DD) for historical entries

class ShopSaleCreate(BaseModel):
    customer_name: Optional[str] = "Walk-in"
    product: str = "milk"
    quantity: Optional[float] = 0
    rate: Optional[float] = 0
    direct_amount: Optional[float] = None
    notes: Optional[str] = ""
    is_udhar: bool = False
    walkin_customer_id: Optional[str] = None

class WalkinCustomerCreate(BaseModel):
    name: str
    phone: str

class UdharPaymentCreate(BaseModel):
    walkin_customer_id: str
    amount: float
    notes: Optional[str] = ""

# Bulk Order Models
class BulkOrderCreate(BaseModel):
    customer_name: str
    customer_phone: str
    customer_type: str = "hotel"  # hotel, caterer, halwai, other
    product: str = "milk"
    quantity: float = 0
    rate: float = 0
    direct_amount: Optional[float] = None
    delivery_address: Optional[str] = ""
    notes: Optional[str] = ""
    is_recurring: bool = False
    status: str = "pending"  # pending, delivered, cancelled


class SaleResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    customer_id: str
    customer_name: str
    product: str
    quantity: float
    rate: float
    amount: float
    date: str
    created_at: str

# Product/Inventory Models
class ProductCreate(BaseModel):
    name: str
    unit: str  # kg, liter, piece
    stock: float = 0
    min_stock: float = 0
    rate: float = 0

class ProductResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    unit: str
    stock: float
    min_stock: float
    rate: float
    updated_at: str

class StockUpdate(BaseModel):
    product_id: str
    quantity: float
    type: str  # in, out
    notes: Optional[str] = ""

# Expense Models
class ExpenseCreate(BaseModel):
    category: str  # salary, transport, electricity, maintenance, other
    amount: float
    description: Optional[str] = ""
    payment_mode: str = "cash"

class ExpenseResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    category: str
    amount: float
    description: str
    payment_mode: str
    date: str
    created_at: str

# Branch Models
class BranchCreate(BaseModel):
    name: str
    code: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    manager_name: Optional[str] = ""

class BranchResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    code: str
    address: str
    phone: str
    manager_name: str
    is_active: bool
    created_at: str

# Bulk Upload Models
class BulkCollectionEntry(BaseModel):
    farmer_phone: str
    shift: str
    quantity: float
    fat: float
    snf: Optional[float] = None

class BulkCollectionUpload(BaseModel):
    entries: List[BulkCollectionEntry]
    branch_id: Optional[str] = None

# Dashboard Models
class DashboardStats(BaseModel):
    total_farmers: int
    active_farmers: int
    today_milk_quantity: float
    today_milk_amount: float
    today_morning_quantity: float
    today_evening_quantity: float
    avg_fat: float
    avg_snf: float
    total_pending_payments: float
    collections_count: int


# ==================== DAIRY PLANT MODELS ====================

# Dairy Plant (Sabar Dairy etc.)
class DairyPlantCreate(BaseModel):
    name: str
    code: Optional[str] = ""
    address: Optional[str] = ""
    phone: Optional[str] = ""
    contact_person: Optional[str] = ""

class DairyPlantUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    contact_person: Optional[str] = None

class DairyPlantResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    code: str
    address: str
    phone: str
    contact_person: str
    total_milk_supplied: float
    total_amount: float
    total_paid: float
    balance: float
    is_active: bool
    created_at: str

# Dispatch Entry (Daily milk dispatch to dairy plant)
class DispatchDeduction(BaseModel):
    type: str  # transport, quality_penalty, commission, testing_charges, other
    amount: float
    notes: Optional[str] = ""

class DispatchCreate(BaseModel):
    dairy_plant_id: str
    date: Optional[str] = None
    tanker_number: Optional[str] = ""
    quantity_kg: float
    avg_fat: float
    avg_snf: float
    clr: Optional[float] = None
    rate_per_kg: float
    deductions: Optional[List[DispatchDeduction]] = []
    notes: Optional[str] = ""

class DispatchResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    dairy_plant_id: str
    dairy_plant_name: str
    date: str
    tanker_number: str
    quantity_kg: float
    avg_fat: float
    avg_snf: float
    clr: Optional[float] = None
    rate_per_kg: float
    gross_amount: float
    deductions: list
    total_deduction: float
    net_receivable: float
    notes: str
    # Slip matching fields
    slip_fat: Optional[float] = None
    slip_snf: Optional[float] = None
    slip_amount: Optional[float] = None
    slip_deductions: Optional[float] = None
    fat_difference: Optional[float] = None
    amount_difference: Optional[float] = None
    slip_matched: bool = False
    created_at: str

# Dispatch Slip Matching
class SlipMatchCreate(BaseModel):
    slip_fat: float
    slip_snf: float
    slip_amount: float
    slip_deductions: Optional[float] = 0

# Dairy Payment
class DairyPaymentCreate(BaseModel):
    dairy_plant_id: str
    amount: float
    payment_mode: str  # cash, upi, bank, cheque
    reference_number: Optional[str] = ""
    notes: Optional[str] = ""

class DairyPaymentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    dairy_plant_id: str
    dairy_plant_name: str
    amount: float
    payment_mode: str
    reference_number: str
    notes: str
    date: str
    created_at: str


# ==================== AUTH UTILITIES ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_access_token(user_id: str, email: str, role: str) -> str:
    expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
        
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ==================== RATE CALCULATION ====================

async def get_milk_rate(fat: float, snf: float) -> float:
    """Calculate milk rate based on fat and SNF using default rate chart"""
    rate_chart = await db.rate_charts.find_one({"is_default": True}, {"_id": 0})
    
    if not rate_chart or not rate_chart.get("entries"):
        # Default formula: Rate = Fat * 6 + SNF * 2 (base formula for Indian dairy)
        return round(fat * 6 + snf * 2, 2)
    
    # Find closest matching rate from chart
    entries = rate_chart["entries"]
    closest_rate = None
    min_diff = float('inf')
    
    for entry in entries:
        diff = abs(entry["fat"] - fat) + abs(entry["snf"] - snf)
        if diff < min_diff:
            min_diff = diff
            closest_rate = entry["rate"]
    
    return closest_rate if closest_rate else round(fat * 6 + snf * 2, 2)

def calculate_snf(fat: float) -> float:
    """Calculate SNF from Fat using standard formula"""
    # Standard formula: SNF = 8.5 + (Fat / 4)
    return round(8.5 + (fat / 4), 2)

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate, current_user: dict = Depends(get_current_user)):
    # Only admin can create new users
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create users")
    
    existing = await db.users.find_one({"email": user.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    user_doc = {
        "id": user_id,
        "name": user.name,
        "email": user.email,
        "phone": user.phone,
        "password": hash_password(user.password),
        "role": user.role,
        "created_at": now,
        "is_active": True
    }
    
    await db.users.insert_one(user_doc)
    
    token = create_access_token(user_id, user.email, user.role)
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user_id,
            name=user.name,
            email=user.email,
            phone=user.phone,
            role=user.role,
            created_at=now
        )
    )

# ==================== ADMIN USER MANAGEMENT ====================

@api_router.get("/users")
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view users")
    users = await db.users.find({}, {"_id": 0, "password": 0}).sort("created_at", -1).to_list(1000)
    return users

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update users")
    
    allowed_fields = {"name", "phone", "role", "is_active"}
    update_data = {k: v for k, v in updates.items() if k in allowed_fields}
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return updated

@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, body: dict, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can reset passwords")
    
    new_password = body.get("password")
    if not new_password or len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    result = await db.users.update_one({"id": user_id}, {"$set": {"password": hash_password(new_password)}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "Password reset successful"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete users")
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted"}

@api_router.post("/admin/login", response_model=TokenResponse)
async def admin_login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access only")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account is disabled")
    
    token = create_access_token(user["id"], user["email"], user["role"])
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user["id"],
            name=user["name"],
            email=user["email"],
            phone=user["phone"],
            role=user["role"],
            created_at=user["created_at"]
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account is disabled")
    
    token = create_access_token(user["id"], user["email"], user["role"])
    
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user["id"],
            name=user["name"],
            email=user["email"],
            phone=user["phone"],
            role=user["role"],
            created_at=user["created_at"]
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        id=current_user["id"],
        name=current_user["name"],
        email=current_user["email"],
        phone=current_user["phone"],
        role=current_user["role"],
        created_at=current_user["created_at"]
    )

# ==================== FARMER ROUTES ====================

@api_router.post("/farmers", response_model=FarmerResponse)
async def create_farmer(farmer: FarmerCreate, current_user: dict = Depends(get_current_user)):
    farmer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    farmer_doc = {
        "id": farmer_id,
        "name": farmer.name,
        "phone": farmer.phone,
        "address": farmer.address or "",
        "village": farmer.village or "",
        "bank_account": farmer.bank_account or "",
        "ifsc_code": farmer.ifsc_code or "",
        "aadhar_number": farmer.aadhar_number or "",
        "milk_type": farmer.milk_type or "cow",
        "fixed_rate": farmer.fixed_rate,
        "cow_rate": farmer.cow_rate,
        "buffalo_rate": farmer.buffalo_rate,
        "total_milk": 0.0,
        "total_due": 0.0,
        "total_paid": 0.0,
        "balance": 0.0,
        "created_at": now,
        "is_active": True
    }
    
    await db.farmers.insert_one(farmer_doc)
    
    return FarmerResponse(**farmer_doc)

@api_router.get("/farmers", response_model=List[FarmerResponse])
async def get_farmers(
    search: Optional[str] = None,
    is_active: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"village": {"$regex": search, "$options": "i"}}
        ]
    if is_active is not None:
        query["is_active"] = is_active
    
    farmers = await db.farmers.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return [FarmerResponse(**f) for f in farmers]

@api_router.get("/farmers/{farmer_id}", response_model=FarmerResponse)
async def get_farmer(farmer_id: str, current_user: dict = Depends(get_current_user)):
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    return FarmerResponse(**farmer)

@api_router.put("/farmers/{farmer_id}", response_model=FarmerResponse)
async def update_farmer(
    farmer_id: str,
    updates: FarmerUpdate,
    current_user: dict = Depends(get_current_user)
):
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.farmers.update_one({"id": farmer_id}, {"$set": update_data})
    
    updated_farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    return FarmerResponse(**updated_farmer)

@api_router.delete("/farmers/{farmer_id}")
async def delete_farmer(farmer_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.farmers.delete_one({"id": farmer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Farmer not found")
    return {"message": "Farmer deleted successfully"}

@api_router.get("/farmers/{farmer_id}/ledger")
async def get_farmer_ledger(
    farmer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    # Get collections
    collection_query = {"farmer_id": farmer_id}
    if start_date:
        collection_query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in collection_query:
            collection_query["date"]["$lte"] = end_date
        else:
            collection_query["date"] = {"$lte": end_date}
    
    collections = await db.milk_collections.find(collection_query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Get payments
    payment_query = {"farmer_id": farmer_id}
    if start_date:
        payment_query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in payment_query:
            payment_query["date"]["$lte"] = end_date
        else:
            payment_query["date"] = {"$lte": end_date}
    
    payments = await db.payments.find(payment_query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    return {
        "farmer": FarmerResponse(**farmer),
        "collections": collections,
        "payments": payments,
        "summary": {
            "total_milk": farmer["total_milk"],
            "total_due": farmer["total_due"],
            "total_paid": farmer["total_paid"],
            "balance": farmer["balance"]
        }
    }

# ==================== MILK COLLECTION ROUTES ====================

@api_router.post("/collections", response_model=MilkCollectionResponse)
async def create_collection(
    collection: MilkCollectionCreate,
    current_user: dict = Depends(get_current_user)
):
    # Get farmer
    farmer = await db.farmers.find_one({"id": collection.farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    # Use provided date or default to today
    date_str = collection.date if collection.date else datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Duplicate entry protection - check if entry already exists for this farmer, date, and shift
    existing = await db.milk_collections.find_one({
        "farmer_id": collection.farmer_id,
        "date": date_str,
        "shift": collection.shift
    }, {"_id": 0})
    
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"Entry already exists for this farmer in {collection.shift} shift on {date_str}. Delete existing entry first."
        )
    
    # Calculate SNF if not provided
    snf = collection.snf if collection.snf else calculate_snf(collection.fat)
    
    # Use provided rate override, or calculate from farmer/rate chart
    milk_type = collection.milk_type or farmer.get("milk_type", "cow")
    rate = None
    if collection.rate and collection.rate > 0:
        rate = collection.rate
    else:
        if milk_type == "buffalo" and farmer.get("buffalo_rate") and farmer["buffalo_rate"] > 0:
            rate = farmer["buffalo_rate"]
        elif milk_type == "cow" and farmer.get("cow_rate") and farmer["cow_rate"] > 0:
            rate = farmer["cow_rate"]
        elif farmer.get("fixed_rate") and farmer["fixed_rate"] > 0:
            rate = farmer["fixed_rate"]
        if not rate:
            rate = await get_milk_rate(collection.fat, snf)
    
    # Calculate amount
    amount = round(collection.quantity * rate, 2)
    
    collection_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    collection_doc = {
        "id": collection_id,
        "farmer_id": collection.farmer_id,
        "farmer_name": farmer["name"],
        "shift": collection.shift,
        "quantity": collection.quantity,
        "fat": collection.fat,
        "snf": snf,
        "rate": rate,
        "amount": amount,
        "milk_type": milk_type,
        "date": date_str,
        "created_at": now.isoformat()
    }
    
    await db.milk_collections.insert_one(collection_doc)
    
    # Update farmer totals
    await db.farmers.update_one(
        {"id": collection.farmer_id},
        {
            "$inc": {
                "total_milk": collection.quantity,
                "total_due": amount,
                "balance": amount
            }
        }
    )
    
    # Send SMS notification (async, don't block on failure)
    try:
        send_collection_sms(
            farmer_name=farmer["name"],
            farmer_phone=farmer["phone"],
            quantity=collection.quantity,
            fat=collection.fat,
            rate=rate,
            amount=amount,
            shift=collection.shift
        )
    except Exception as e:
        logger.warning(f"Failed to send collection SMS: {e}")
    
    return MilkCollectionResponse(**collection_doc)

@api_router.get("/collections", response_model=List[MilkCollectionResponse])
async def get_collections(
    date: Optional[str] = None,
    farmer_id: Optional[str] = None,
    shift: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if date:
        query["date"] = date
    if farmer_id:
        query["farmer_id"] = farmer_id
    if shift:
        query["shift"] = shift
    
    collections = await db.milk_collections.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [MilkCollectionResponse(**c) for c in collections]

@api_router.get("/collections/today", response_model=List[MilkCollectionResponse])
async def get_today_collections(
    shift: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    query = {"date": today}
    if shift:
        query["shift"] = shift
    
    collections = await db.milk_collections.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [MilkCollectionResponse(**c) for c in collections]

@api_router.delete("/collections/{collection_id}")
async def delete_collection(collection_id: str, current_user: dict = Depends(get_current_user)):
    collection = await db.milk_collections.find_one({"id": collection_id}, {"_id": 0})
    if not collection:
        raise HTTPException(status_code=404, detail="Collection not found")
    
    # Revert farmer totals
    await db.farmers.update_one(
        {"id": collection["farmer_id"]},
        {
            "$inc": {
                "total_milk": -collection["quantity"],
                "total_due": -collection["amount"],
                "balance": -collection["amount"]
            }
        }
    )
    
    await db.milk_collections.delete_one({"id": collection_id})
    return {"message": "Collection deleted successfully"}

# ==================== RATE CHART ROUTES ====================

@api_router.post("/rate-charts", response_model=RateChartResponse)
async def create_rate_chart(
    rate_chart: RateChartCreate,
    current_user: dict = Depends(get_current_user)
):
    chart_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    # If this is default, unset other defaults
    if rate_chart.is_default:
        await db.rate_charts.update_many({}, {"$set": {"is_default": False}})
    
    entries = [e.model_dump() for e in rate_chart.entries]
    
    chart_doc = {
        "id": chart_id,
        "name": rate_chart.name,
        "entries": entries,
        "is_default": rate_chart.is_default,
        "created_at": now,
        "updated_at": now
    }
    
    await db.rate_charts.insert_one(chart_doc)
    return RateChartResponse(**chart_doc)

@api_router.get("/rate-charts", response_model=List[RateChartResponse])
async def get_rate_charts(current_user: dict = Depends(get_current_user)):
    charts = await db.rate_charts.find({}, {"_id": 0}).to_list(100)
    return [RateChartResponse(**c) for c in charts]

@api_router.get("/rate-charts/default", response_model=RateChartResponse)
async def get_default_rate_chart(current_user: dict = Depends(get_current_user)):
    chart = await db.rate_charts.find_one({"is_default": True}, {"_id": 0})
    if not chart:
        raise HTTPException(status_code=404, detail="No default rate chart found")
    return RateChartResponse(**chart)

@api_router.put("/rate-charts/{chart_id}", response_model=RateChartResponse)
async def update_rate_chart(
    chart_id: str,
    rate_chart: RateChartCreate,
    current_user: dict = Depends(get_current_user)
):
    existing = await db.rate_charts.find_one({"id": chart_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Rate chart not found")
    
    if rate_chart.is_default:
        await db.rate_charts.update_many({}, {"$set": {"is_default": False}})
    
    entries = [e.model_dump() for e in rate_chart.entries]
    now = datetime.now(timezone.utc).isoformat()
    
    await db.rate_charts.update_one(
        {"id": chart_id},
        {
            "$set": {
                "name": rate_chart.name,
                "entries": entries,
                "is_default": rate_chart.is_default,
                "updated_at": now
            }
        }
    )
    
    updated = await db.rate_charts.find_one({"id": chart_id}, {"_id": 0})
    return RateChartResponse(**updated)

@api_router.delete("/rate-charts/{chart_id}")
async def delete_rate_chart(chart_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.rate_charts.delete_one({"id": chart_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Rate chart not found")
    return {"message": "Rate chart deleted successfully"}

@api_router.post("/rate-charts/calculate-rate")
async def calculate_rate(
    fat: float,
    snf: Optional[float] = None,
    current_user: dict = Depends(get_current_user)
):
    if snf is None:
        snf = calculate_snf(fat)
    rate = await get_milk_rate(fat, snf)
    return {"fat": fat, "snf": snf, "rate": rate}

# ==================== PAYMENT ROUTES ====================

@api_router.post("/payments", response_model=PaymentResponse)
async def create_payment(
    payment: PaymentCreate,
    current_user: dict = Depends(get_current_user)
):
    farmer = await db.farmers.find_one({"id": payment.farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%Y-%m-%d")
    
    payment_doc = {
        "id": payment_id,
        "farmer_id": payment.farmer_id,
        "farmer_name": farmer["name"],
        "amount": payment.amount,
        "payment_mode": payment.payment_mode,
        "payment_type": payment.payment_type,
        "notes": payment.notes or "",
        "date": date_str,
        "created_at": now.isoformat()
    }
    
    await db.payments.insert_one(payment_doc)
    
    # Update farmer totals based on payment type
    if payment.payment_type == "advance":
        # Advance increases balance (farmer owes us)
        await db.farmers.update_one(
            {"id": payment.farmer_id},
            {"$inc": {"total_paid": payment.amount, "balance": payment.amount}}
        )
    elif payment.payment_type == "deduction":
        # Deduction reduces from due amount
        await db.farmers.update_one(
            {"id": payment.farmer_id},
            {"$inc": {"total_due": -payment.amount, "balance": -payment.amount}}
        )
    else:
        # Normal payment
        await db.farmers.update_one(
            {"id": payment.farmer_id},
            {"$inc": {"total_paid": payment.amount, "balance": -payment.amount}}
        )
    
    # Calculate new balance and send SMS
    new_balance = farmer["balance"] - payment.amount
    try:
        send_payment_sms(
            farmer_name=farmer["name"],
            farmer_phone=farmer["phone"],
            amount=payment.amount,
            payment_mode=payment.payment_mode,
            new_balance=new_balance
        )
    except Exception as e:
        logger.warning(f"Failed to send payment SMS: {e}")
    
    return PaymentResponse(**payment_doc)

@api_router.get("/payments", response_model=List[PaymentResponse])
async def get_payments(
    farmer_id: Optional[str] = None,
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if farmer_id:
        query["farmer_id"] = farmer_id
    if date:
        query["date"] = date
    
    payments = await db.payments.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [PaymentResponse(**p) for p in payments]

@api_router.delete("/payments/{payment_id}")
async def delete_payment(payment_id: str, current_user: dict = Depends(get_current_user)):
    payment = await db.payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    # Revert farmer totals
    await db.farmers.update_one(
        {"id": payment["farmer_id"]},
        {
            "$inc": {
                "total_paid": -payment["amount"],
                "balance": payment["amount"]
            }
        }
    )
    
    await db.payments.delete_one({"id": payment_id})
    return {"message": "Payment deleted successfully"}

# ==================== CUSTOMER ROUTES ====================

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    customer_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    customer_doc = {
        "id": customer_id,
        "name": customer.name,
        "phone": customer.phone,
        "address": customer.address or "",
        "customer_type": customer.customer_type,
        "gst_number": customer.gst_number or "",
        "total_purchase": 0.0,
        "total_paid": 0.0,
        "balance": 0.0,
        "created_at": now,
        "is_active": True
    }
    
    await db.customers.insert_one(customer_doc)
    return CustomerResponse(**customer_doc)

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(
    search: Optional[str] = None,
    customer_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}}
        ]
    if customer_type:
        query["customer_type"] = customer_type
    
    customers = await db.customers.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return [CustomerResponse(**c) for c in customers]

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return CustomerResponse(**customer)

@api_router.put("/customers/{customer_id}", response_model=CustomerResponse)
async def update_customer(
    customer_id: str,
    updates: CustomerUpdate,
    current_user: dict = Depends(get_current_user)
):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.customers.update_one({"id": customer_id}, {"$set": update_data})
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return CustomerResponse(**updated)



# ==================== SALES ROUTES ====================

@api_router.post("/sales", response_model=SaleResponse)
async def create_sale(sale: SaleCreate, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": sale.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    sale_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    date_str = sale.date if sale.date else now.strftime("%Y-%m-%d")
    
    if sale.direct_amount and sale.direct_amount > 0:
        amount = round(sale.direct_amount, 2)
        quantity = sale.quantity or 0
        rate = sale.rate or 0
    else:
        quantity = sale.quantity or 0
        rate = sale.rate or 0
        amount = round(quantity * rate, 2)
    
    sale_doc = {
        "id": sale_id,
        "customer_id": sale.customer_id,
        "customer_name": customer["name"],
        "product": sale.product,
        "quantity": quantity,
        "rate": rate,
        "amount": amount,
        "notes": sale.notes or "",
        "date": date_str,
        "created_at": now.isoformat()
    }
    
    await db.sales.insert_one(sale_doc)
    
    # Update customer totals
    await db.customers.update_one(
        {"id": sale.customer_id},
        {"$inc": {"total_purchase": amount, "balance": amount}}
    )
    
    # Update product stock if exists
    await db.products.update_one(
        {"name": sale.product},
        {"$inc": {"stock": -sale.quantity}}
    )
    
    return SaleResponse(**sale_doc)

@api_router.get("/sales", response_model=List[SaleResponse])
async def get_sales(
    date: Optional[str] = None,
    customer_id: Optional[str] = None,
    product: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if date:
        query["date"] = date
    if customer_id:
        query["customer_id"] = customer_id
    if product:
        query["product"] = product
    
    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [SaleResponse(**s) for s in sales]

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, current_user: dict = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    # Revert customer totals
    await db.customers.update_one(
        {"id": sale["customer_id"]},
        {"$inc": {"total_purchase": -sale["amount"], "balance": -sale["amount"]}}
    )
    
    await db.sales.delete_one({"id": sale_id})
    return {"message": "Sale deleted successfully"}

@api_router.get("/sales/today")
async def get_today_sales(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    sales = await db.sales.find({"date": today}, {"_id": 0}).to_list(1000)
    
    total_amount = sum(s["amount"] for s in sales)
    by_product = {}
    for s in sales:
        if s["product"] not in by_product:
            by_product[s["product"]] = {"quantity": 0, "amount": 0}
        by_product[s["product"]]["quantity"] += s["quantity"]
        by_product[s["product"]]["amount"] += s["amount"]
    
    return {
        "date": today,
        "total_sales": len(sales),
        "total_amount": round(total_amount, 2),
        "by_product": by_product,
        "sales": sales
    }

@api_router.post("/sales/shop")
async def create_shop_sale(sale: ShopSaleCreate, current_user: dict = Depends(get_current_user)):
    """Quick shop/counter milk sale - supports udhar (credit) and direct amount"""
    sale_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%Y-%m-%d")
    
    if sale.direct_amount and sale.direct_amount > 0:
        amount = round(sale.direct_amount, 2)
        quantity = sale.quantity or 0
        rate = sale.rate or 0
    else:
        quantity = sale.quantity or 0
        rate = sale.rate or 0
        amount = round(quantity * rate, 2)
    
    customer_name = sale.customer_name or "Walk-in"
    
    # If udhar, validate walk-in customer exists
    if sale.is_udhar and sale.walkin_customer_id:
        wc = await db.walkin_customers.find_one({"id": sale.walkin_customer_id}, {"_id": 0})
        if wc:
            customer_name = wc["name"]
    
    sale_doc = {
        "id": sale_id,
        "customer_id": sale.walkin_customer_id or "shop_walk_in",
        "customer_name": customer_name,
        "product": sale.product,
        "quantity": sale.quantity,
        "rate": sale.rate,
        "amount": amount,
        "notes": sale.notes or "",
        "is_shop_sale": True,
        "is_udhar": sale.is_udhar,
        "date": date_str,
        "created_at": now.isoformat()
    }
    await db.sales.insert_one(sale_doc)
    del sale_doc["_id"]
    
    # Update walk-in customer balance if udhar
    if sale.is_udhar and sale.walkin_customer_id:
        await db.walkin_customers.update_one(
            {"id": sale.walkin_customer_id},
            {"$inc": {"pending_amount": amount}}
        )
    
    return sale_doc

# ==================== WALK-IN CUSTOMER & UDHAR ROUTES ====================

@api_router.post("/walkin-customers")
async def create_walkin_customer(data: WalkinCustomerCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.walkin_customers.find_one({"phone": data.phone}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Customer with this phone already exists")
    
    cust_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": cust_id,
        "name": data.name,
        "phone": data.phone,
        "pending_amount": 0,
        "total_purchases": 0,
        "total_paid": 0,
        "created_at": now,
        "is_active": True
    }
    await db.walkin_customers.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.get("/walkin-customers")
async def get_walkin_customers(current_user: dict = Depends(get_current_user)):
    customers = await db.walkin_customers.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return customers

@api_router.get("/walkin-customers/{customer_id}")
async def get_walkin_customer_detail(customer_id: str, current_user: dict = Depends(get_current_user)):
    customer = await db.walkin_customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get all udhar sales
    sales = await db.sales.find(
        {"customer_id": customer_id, "is_udhar": True},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    # Get all payments
    payments = await db.udhar_payments.find(
        {"walkin_customer_id": customer_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    return {
        "customer": customer,
        "sales": sales,
        "payments": payments
    }

@api_router.post("/udhar-payments")
async def record_udhar_payment(payment: UdharPaymentCreate, current_user: dict = Depends(get_current_user)):
    customer = await db.walkin_customers.find_one({"id": payment.walkin_customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    doc = {
        "id": payment_id,
        "walkin_customer_id": payment.walkin_customer_id,
        "amount": payment.amount,
        "notes": payment.notes or "",
        "date": now.strftime("%Y-%m-%d"),
        "created_at": now.isoformat()
    }
    await db.udhar_payments.insert_one(doc)
    del doc["_id"]
    
    # Update customer balance
    await db.walkin_customers.update_one(
        {"id": payment.walkin_customer_id},
        {
            "$inc": {"pending_amount": -payment.amount, "total_paid": payment.amount}
        }
    )
    
    return doc


# ==================== BILLING ROUTES ====================

@api_router.get("/billing/farmer/{farmer_id}")
async def get_farmer_billing(farmer_id: str, start_date: str, end_date: str, current_user: dict = Depends(get_current_user)):
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    collections = await db.milk_collections.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0}
    ).sort("date", 1).to_list(5000)
    
    payments = await db.payments.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}},
        {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    total_quantity = sum(c.get("quantity", 0) for c in collections)
    total_amount = sum(c.get("amount", 0) for c in collections)
    total_paid = sum(p.get("amount", 0) for p in payments)
    
    return {
        "farmer": farmer,
        "collections": collections,
        "payments": payments,
        "summary": {
            "total_quantity": round(total_quantity, 2),
            "total_amount": round(total_amount, 2),
            "total_paid": round(total_paid, 2),
            "balance_due": round(total_amount - total_paid, 2),
            "total_entries": len(collections),
            "start_date": start_date,
            "end_date": end_date
        }
    }

@api_router.get("/billing/customer/{customer_id}")
async def get_customer_billing(customer_id: str, start_date: str, end_date: str, current_user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    sales = await db.sales.find(
        {"customer_id": customer_id, "date": {"$gte": start_date, "$lte": end_date}, "is_shop_sale": {"$ne": True}},
        {"_id": 0}
    ).sort("date", 1).to_list(5000)
    
    total_quantity = sum(s.get("quantity", 0) for s in sales)
    total_amount = sum(s.get("amount", 0) for s in sales)
    
    return {
        "customer": customer,
        "sales": sales,
        "summary": {
            "total_quantity": round(total_quantity, 2),
            "total_amount": round(total_amount, 2),
            "total_entries": len(sales),
            "start_date": start_date,
            "end_date": end_date
        }
    }

# ==================== BULK ORDER ROUTES ====================

@api_router.post("/bulk-orders")
async def create_bulk_order(order: BulkOrderCreate, current_user: dict = Depends(get_current_user)):
    order_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    if order.direct_amount and order.direct_amount > 0:
        amount = round(order.direct_amount, 2)
    else:
        amount = round(order.quantity * order.rate, 2)
    
    doc = {
        "id": order_id,
        "customer_name": order.customer_name,
        "customer_phone": order.customer_phone,
        "customer_type": order.customer_type,
        "product": order.product,
        "quantity": order.quantity,
        "rate": order.rate,
        "amount": amount,
        "delivery_address": order.delivery_address or "",
        "notes": order.notes or "",
        "is_recurring": order.is_recurring,
        "status": order.status,
        "date": now.strftime("%Y-%m-%d"),
        "created_at": now.isoformat()
    }
    await db.bulk_orders.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.get("/bulk-orders")
async def get_bulk_orders(date: Optional[str] = None, status: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if date:
        query["date"] = date
    if status:
        query["status"] = status
    orders = await db.bulk_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return orders

@api_router.put("/bulk-orders/{order_id}")
async def update_bulk_order(order_id: str, updates: dict, current_user: dict = Depends(get_current_user)):
    allowed = {"status", "quantity", "rate", "amount", "direct_amount", "notes", "delivery_address", "product"}
    update_data = {k: v for k, v in updates.items() if k in allowed}
    
    # Recalculate amount if needed
    if "direct_amount" in update_data and update_data["direct_amount"]:
        update_data["amount"] = round(update_data["direct_amount"], 2)
    elif "quantity" in update_data and "rate" in update_data:
        update_data["amount"] = round(update_data["quantity"] * update_data["rate"], 2)
    
    if "direct_amount" in update_data:
        del update_data["direct_amount"]
    
    result = await db.bulk_orders.update_one({"id": order_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    
    updated = await db.bulk_orders.find_one({"id": order_id}, {"_id": 0})
    return updated

@api_router.delete("/bulk-orders/{order_id}")
async def delete_bulk_order(order_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.bulk_orders.delete_one({"id": order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Order not found")
    return {"message": "Order deleted"}

@api_router.get("/bulk-order-customers")
async def get_bulk_order_customers(current_user: dict = Depends(get_current_user)):
    """Get unique customers from bulk orders for search/autocomplete"""
    pipeline = [
        {"$group": {
            "_id": {"name": "$customer_name", "phone": "$customer_phone", "type": "$customer_type"},
            "total_orders": {"$sum": 1},
            "total_amount": {"$sum": "$amount"},
            "last_order": {"$max": "$created_at"}
        }},
        {"$sort": {"last_order": -1}},
        {"$project": {
            "_id": 0,
            "name": "$_id.name",
            "phone": "$_id.phone",
            "customer_type": "$_id.type",
            "total_orders": 1,
            "total_amount": 1,
            "last_order": 1
        }}
    ]
    customers = await db.bulk_orders.aggregate(pipeline).to_list(500)
    return customers


# ==================== CUSTOMER BILLING ROUTES ====================

@api_router.get("/customers/{customer_id}/sales")
async def get_customer_sales(
    customer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all sales for a customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    query = {"customer_id": customer_id}
    if start_date:
        query.setdefault("date", {})["$gte"] = start_date
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    
    sales = await db.sales.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    payments = await db.payments.find({"farmer_id": customer_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    
    return {
        "customer": customer,
        "sales": sales,
        "payments": payments,
        "summary": {
            "total_purchase": customer.get("total_purchase", 0),
            "total_paid": customer.get("total_paid", 0),
            "balance": customer.get("balance", 0),
        }
    }

@api_router.post("/customers/{customer_id}/payment")
async def record_customer_payment(
    customer_id: str,
    payment: dict,
    current_user: dict = Depends(get_current_user)
):
    """Record a payment from a customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    amount = float(payment.get("amount", 0))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    
    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    
    payment_doc = {
        "id": payment_id,
        "farmer_id": customer_id,
        "farmer_name": customer["name"],
        "amount": amount,
        "payment_mode": payment.get("payment_mode", "cash"),
        "payment_type": "payment",
        "notes": payment.get("notes", ""),
        "date": now.strftime("%Y-%m-%d"),
        "created_at": now.isoformat()
    }
    
    await db.payments.insert_one(payment_doc)
    await db.customers.update_one(
        {"id": customer_id},
        {"$inc": {"total_paid": amount, "balance": -amount}}
    )
    
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    return {"payment": payment_doc, "customer": updated}

@api_router.get("/bills/customer/thermal/{customer_id}", response_class=HTMLResponse)
async def customer_thermal_bill(customer_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Generate thermal printer bill for customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    sales = await db.sales.find(
        {"customer_id": customer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    payments = await db.payments.find(
        {"farmer_id": customer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    dairy_phone = settings.get("phone", "")
    
    total_amount = sum(s["amount"] for s in sales)
    total_paid = sum(p["amount"] for p in payments)
    balance = total_amount - total_paid
    
    rows = ""
    for s in sales:
        rows += f"<tr><td>{s['date'][5:]}</td><td>{s['product']}</td><td>{s['quantity']}</td><td>{s['rate']}</td><td style='text-align:right'>{s['amount']:.0f}</td></tr>\n"
    
    pay_rows = ""
    for p in payments:
        pay_rows += f"<tr><td>{p['date'][5:]}</td><td>{p['payment_mode']}</td><td style='text-align:right'>{p['amount']:.0f}</td></tr>\n"
    
    ctype = "Wholesale" if customer.get("customer_type") == "wholesale" else "Retail"
    
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=58mm">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Courier New',monospace;font-size:11px;width:58mm;padding:2mm;color:#000}}
.center{{text-align:center}}.bold{{font-weight:bold}}.line{{border-top:1px dashed #000;margin:3px 0}}
table{{width:100%;border-collapse:collapse}}
td,th{{padding:1px 2px;font-size:10px;vertical-align:top}}
th{{text-align:left;border-bottom:1px solid #000}}
.right{{text-align:right}}.big{{font-size:14px}}
@media print{{@page{{size:58mm auto;margin:0}}body{{width:58mm}}}}
</style></head><body>
<div class="center bold big">{dairy_name}</div>
<div class="center" style="font-size:9px">{dairy_phone}</div>
<div class="line"></div>
<div class="bold">{customer['name']} ({ctype})</div>
<div style="font-size:9px">Ph: {customer['phone']}</div>
{f'<div style="font-size:9px">GST: {customer["gst_number"]}</div>' if customer.get("gst_number") else ''}
<div style="font-size:9px">{start_date} to {end_date}</div>
<div class="line"></div>
<table><tr><th>Date</th><th>Item</th><th>Qty</th><th>Rate</th><th class="right">Amt</th></tr>
{rows}</table>
<div class="line"></div>
<table>
<tr><td class="bold">Total:</td><td class="right bold">Rs.{total_amount:.0f}</td></tr>
<tr><td class="bold">Paid:</td><td class="right bold">Rs.{total_paid:.0f}</td></tr>
<tr><td class="bold">Balance:</td><td class="right bold">Rs.{balance:.0f}</td></tr>
</table>
{f'<div class="line"></div><div class="bold" style="font-size:9px">Payments:</div><table><tr><th>Date</th><th>Mode</th><th class="right">Amt</th></tr>{pay_rows}</table>' if payments else ''}
<div class="line"></div>
<div class="center" style="font-size:9px;margin-top:3px">Thank You / धन्यवाद</div>
<div class="center" style="font-size:8px">Printed: {datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M")}</div>
</body></html>"""
    return HTMLResponse(content=html)

@api_router.get("/bills/customer/a4/{customer_id}", response_class=HTMLResponse)
async def customer_a4_invoice(customer_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Generate A4 invoice for customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    sales = await db.sales.find(
        {"customer_id": customer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    payments = await db.payments.find(
        {"farmer_id": customer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    dairy_phone = settings.get("phone", "")
    dairy_address = settings.get("address", "")
    
    total_amount = sum(s["amount"] for s in sales)
    total_paid = sum(p["amount"] for p in payments)
    balance = total_amount - total_paid
    ctype = "Wholesale / थोक" if customer.get("customer_type") == "wholesale" else "Retail / खुदरा"
    invoice_no = f"CINV-{customer_id[:8].upper()}-{datetime.now(timezone.utc).strftime('%Y%m%d')}"
    
    sale_rows = ""
    for i, s in enumerate(sales, 1):
        sale_rows += f"<tr><td>{i}</td><td>{s['date']}</td><td>{s['product'].title()}</td><td>{s['quantity']}</td><td>{s['rate']:.2f}</td><td class='right'>{s['amount']:.2f}</td></tr>"
    
    pay_rows = ""
    for p in payments:
        pay_rows += f"<tr><td>{p['date']}</td><td>{p.get('payment_mode','cash').upper()}</td><td>{p.get('notes','')}</td><td class='right'>{p['amount']:.2f}</td></tr>"
    
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1a1a1a;padding:15mm 20mm;max-width:210mm}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #15803d;padding-bottom:12px;margin-bottom:15px}}
.dairy-name{{font-size:24px;font-weight:bold;color:#15803d}}.dairy-info{{font-size:10px;color:#666;margin-top:4px}}
.invoice-box{{text-align:right}}.invoice-title{{font-size:20px;color:#15803d;font-weight:bold}}.invoice-no{{font-size:11px;color:#666;margin-top:2px}}
.info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:15px;margin-bottom:15px}}
.info-card{{background:#f8faf8;border:1px solid #e5e7eb;border-radius:6px;padding:10px}}
.info-label{{font-size:9px;color:#888;text-transform:uppercase;letter-spacing:0.5px}}.info-value{{font-size:13px;font-weight:600;margin-top:2px}}
table{{width:100%;border-collapse:collapse;margin-bottom:15px}}
th{{background:#15803d;color:white;padding:6px 8px;text-align:left;font-size:10px;text-transform:uppercase}}
td{{padding:5px 8px;border-bottom:1px solid #eee;font-size:11px}}.right{{text-align:right}}
tr:nth-child(even){{background:#f9fafb}}
.summary-grid{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin:15px 0}}
.summary-box{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:10px;text-align:center}}
.summary-box.due{{background:#fef2f2;border-color:#fecaca}}
.summary-label{{font-size:9px;color:#666;text-transform:uppercase}}.summary-value{{font-size:18px;font-weight:bold;color:#15803d;margin-top:2px}}
.summary-box.due .summary-value{{color:#dc2626}}
.footer{{margin-top:20px;border-top:2px solid #15803d;padding-top:10px;display:flex;justify-content:space-between;font-size:10px;color:#666}}
.sig-box{{border-top:1px solid #333;width:150px;text-align:center;padding-top:5px;margin-top:40px;font-size:10px}}
h3{{color:#15803d;font-size:13px;margin-bottom:8px;padding-bottom:4px;border-bottom:1px solid #d1fae5}}
@media print{{@page{{size:A4;margin:10mm}}body{{padding:0}}}}
</style></head><body>
<div class="header">
  <div><div class="dairy-name">{dairy_name}</div><div class="dairy-info">{dairy_address}<br>{dairy_phone}</div></div>
  <div class="invoice-box"><div class="invoice-title">CUSTOMER INVOICE / ग्राहक बिल</div><div class="invoice-no">{invoice_no}</div><div class="invoice-no">{start_date} to {end_date}</div></div>
</div>
<div class="info-grid">
  <div class="info-card"><div class="info-label">Customer / ग्राहक</div><div class="info-value">{customer['name']}</div><div style="font-size:10px;color:#666">Ph: {customer['phone']} | {ctype}</div></div>
  <div class="info-card"><div class="info-label">Address / पता</div><div class="info-value">{customer.get('address','N/A')}</div>{f'<div style="font-size:10px;color:#666">GST: {customer["gst_number"]}</div>' if customer.get("gst_number") else ''}</div>
</div>
<div class="summary-grid">
  <div class="summary-box"><div class="summary-label">Total Purchase / कुल खरीद</div><div class="summary-value">₹{total_amount:.0f}</div></div>
  <div class="summary-box"><div class="summary-label">Total Paid / कुल भुगतान</div><div class="summary-value">₹{total_paid:.0f}</div></div>
  <div class="summary-box {'due' if balance > 0 else ''}"><div class="summary-label">Balance / बकाया</div><div class="summary-value">₹{balance:.0f}</div></div>
</div>
<h3>Sales Details / बिक्री विवरण</h3>
<table><tr><th>#</th><th>Date</th><th>Product</th><th>Qty</th><th>Rate ₹</th><th class="right">Amount ₹</th></tr>
{sale_rows}
<tr style="background:#15803d;color:white;font-weight:bold"><td colspan="5">TOTAL</td><td class="right">₹{total_amount:.2f}</td></tr>
</table>
{f'<h3>Payments / भुगतान</h3><table><tr><th>Date</th><th>Mode</th><th>Notes</th><th class="right">Amount ₹</th></tr>{pay_rows}<tr style="background:#15803d;color:white;font-weight:bold"><td colspan="3">TOTAL PAID</td><td class="right">₹{total_paid:.2f}</td></tr></table>' if payments else ''}
<div style="display:flex;justify-content:flex-end;gap:40px;margin-top:30px">
  <div class="sig-box">Dairy Stamp / डेयरी मुहर</div>
  <div class="sig-box">Authorized Signature / हस्ताक्षर</div>
</div>
<div class="footer"><div>Generated: {datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M UTC")}</div><div>{dairy_name} | {dairy_phone}</div></div>
</body></html>"""
    return HTMLResponse(content=html)

@api_router.get("/share/customer-bill/{customer_id}")
async def share_customer_bill(customer_id: str, current_user: dict = Depends(get_current_user)):
    """Generate WhatsApp share link for customer bill"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    
    msg = f"*{dairy_name}*\nCustomer: {customer['name']}\nTotal Purchase: ₹{customer.get('total_purchase', 0):.0f}\nPaid: ₹{customer.get('total_paid', 0):.0f}\nBalance: ₹{customer.get('balance', 0):.0f}\nThank you!"
    
    import urllib.parse
    phone = customer["phone"].replace("+", "").replace(" ", "")
    if not phone.startswith("91"):
        phone = "91" + phone
    
    whatsapp_link = f"https://wa.me/{phone}?text={urllib.parse.quote(msg)}"
    return {"whatsapp_link": whatsapp_link, "message": msg}

# ==================== INVENTORY/PRODUCT ROUTES ====================

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    product_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    product_doc = {
        "id": product_id,
        "name": product.name,
        "unit": product.unit,
        "stock": product.stock,
        "min_stock": product.min_stock,
        "rate": product.rate,
        "updated_at": now
    }
    
    await db.products.insert_one(product_doc)
    return ProductResponse(**product_doc)

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(current_user: dict = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    return [ProductResponse(**p) for p in products]

@api_router.get("/products/low-stock")
async def get_low_stock_products(current_user: dict = Depends(get_current_user)):
    products = await db.products.find({}, {"_id": 0}).to_list(100)
    low_stock = [p for p in products if p["stock"] <= p["min_stock"]]
    return low_stock

@api_router.post("/products/stock-update")
async def update_stock(update: StockUpdate, current_user: dict = Depends(get_current_user)):
    product = await db.products.find_one({"id": update.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    quantity_change = update.quantity if update.type == "in" else -update.quantity
    now = datetime.now(timezone.utc).isoformat()
    
    await db.products.update_one(
        {"id": update.product_id},
        {"$inc": {"stock": quantity_change}, "$set": {"updated_at": now}}
    )
    
    # Log stock movement
    await db.stock_movements.insert_one({
        "id": str(uuid.uuid4()),
        "product_id": update.product_id,
        "product_name": product["name"],
        "quantity": update.quantity,
        "type": update.type,
        "notes": update.notes or "",
        "created_at": now
    })
    
    return {"message": "Stock updated successfully"}

# ==================== EXPENSE ROUTES ====================

@api_router.post("/expenses", response_model=ExpenseResponse)
async def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    expense_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    date_str = now.strftime("%Y-%m-%d")
    
    expense_doc = {
        "id": expense_id,
        "category": expense.category,
        "amount": expense.amount,
        "description": expense.description or "",
        "payment_mode": expense.payment_mode,
        "date": date_str,
        "created_at": now.isoformat()
    }
    
    await db.expenses.insert_one(expense_doc)
    return ExpenseResponse(**expense_doc)

@api_router.get("/expenses", response_model=List[ExpenseResponse])
async def get_expenses(
    date: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if date:
        query["date"] = date
    if category:
        query["category"] = category
    if start_date and end_date:
        query["date"] = {"$gte": start_date, "$lte": end_date}
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [ExpenseResponse(**e) for e in expenses]

@api_router.get("/expenses/summary")
async def get_expense_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if not start_date:
        today = datetime.now(timezone.utc)
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    expenses = await db.expenses.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(1000)
    
    by_category = {}
    total = 0
    for e in expenses:
        cat = e["category"]
        if cat not in by_category:
            by_category[cat] = 0
        by_category[cat] += e["amount"]
        total += e["amount"]
    
    return {
        "period": {"start": start_date, "end": end_date},
        "total": round(total, 2),
        "by_category": by_category,
        "count": len(expenses)
    }

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

# ==================== BRANCH MANAGEMENT ROUTES ====================

@api_router.post("/branches", response_model=BranchResponse)
async def create_branch(branch: BranchCreate, current_user: dict = Depends(get_current_user)):
    # Check if branch code already exists
    existing = await db.branches.find_one({"code": branch.code}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Branch code already exists")
    
    branch_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    branch_doc = {
        "id": branch_id,
        "name": branch.name,
        "code": branch.code.upper(),
        "address": branch.address or "",
        "phone": branch.phone or "",
        "manager_name": branch.manager_name or "",
        "is_active": True,
        "created_at": now
    }
    
    await db.branches.insert_one(branch_doc)
    return BranchResponse(**branch_doc)

@api_router.get("/branches", response_model=List[BranchResponse])
async def get_branches(current_user: dict = Depends(get_current_user)):
    branches = await db.branches.find({}, {"_id": 0}).sort("name", 1).to_list(100)
    return [BranchResponse(**b) for b in branches]

@api_router.get("/branches/{branch_id}", response_model=BranchResponse)
async def get_branch(branch_id: str, current_user: dict = Depends(get_current_user)):
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    return BranchResponse(**branch)

@api_router.put("/branches/{branch_id}", response_model=BranchResponse)
async def update_branch(branch_id: str, branch: BranchCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    await db.branches.update_one(
        {"id": branch_id},
        {"$set": {
            "name": branch.name,
            "code": branch.code.upper(),
            "address": branch.address or "",
            "phone": branch.phone or "",
            "manager_name": branch.manager_name or ""
        }}
    )
    
    updated = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    return BranchResponse(**updated)

@api_router.delete("/branches/{branch_id}")
async def delete_branch(branch_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.branches.delete_one({"id": branch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"message": "Branch deleted successfully"}

@api_router.get("/branches/{branch_id}/stats")
async def get_branch_stats(branch_id: str, current_user: dict = Depends(get_current_user)):
    """Get statistics for a specific branch"""
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get branch collections
    collections = await db.milk_collections.find(
        {"branch_id": branch_id, "date": today}, {"_id": 0}
    ).to_list(1000)
    
    # Get branch farmers
    farmers_count = await db.farmers.count_documents({"branch_id": branch_id})
    
    total_quantity = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    
    return {
        "branch": branch,
        "today": {
            "collections": len(collections),
            "quantity": round(total_quantity, 2),
            "amount": round(total_amount, 2)
        },
        "total_farmers": farmers_count
    }

# ==================== BULK UPLOAD ROUTES ====================

@api_router.post("/bulk/collections")
async def bulk_upload_collections(
    upload: BulkCollectionUpload,
    current_user: dict = Depends(get_current_user)
):
    """Bulk upload milk collections from CSV/Excel data"""
    results = {
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    for entry in upload.entries:
        try:
            # Find farmer by phone
            farmer = await db.farmers.find_one({"phone": entry.farmer_phone}, {"_id": 0})
            if not farmer:
                results["failed"] += 1
                results["errors"].append(f"Farmer not found: {entry.farmer_phone}")
                continue
            
            # Check for duplicate
            today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            existing = await db.milk_collections.find_one({
                "farmer_id": farmer["id"],
                "date": today,
                "shift": entry.shift
            }, {"_id": 0})
            
            if existing:
                results["failed"] += 1
                results["errors"].append(f"Duplicate entry for {farmer['name']} ({entry.shift})")
                continue
            
            # Calculate SNF and rate
            snf = entry.snf if entry.snf else calculate_snf(entry.fat)
            rate = await get_milk_rate(entry.fat, snf)
            amount = round(entry.quantity * rate, 2)
            
            collection_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc)
            
            collection_doc = {
                "id": collection_id,
                "farmer_id": farmer["id"],
                "farmer_name": farmer["name"],
                "shift": entry.shift,
                "quantity": entry.quantity,
                "fat": entry.fat,
                "snf": snf,
                "rate": rate,
                "amount": amount,
                "date": today,
                "branch_id": upload.branch_id or "",
                "created_at": now.isoformat()
            }
            
            await db.milk_collections.insert_one(collection_doc)
            
            # Update farmer totals
            await db.farmers.update_one(
                {"id": farmer["id"]},
                {"$inc": {"total_milk": entry.quantity, "total_due": amount, "balance": amount}}
            )
            
            results["success"] += 1
            
        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"Error processing {entry.farmer_phone}: {str(e)}")
    
    return results

@api_router.post("/bulk/farmers")
async def bulk_upload_farmers(
    farmers_data: List[FarmerCreate],
    current_user: dict = Depends(get_current_user)
):
    """Bulk upload farmers from CSV/Excel data"""
    results = {
        "success": 0,
        "failed": 0,
        "errors": []
    }
    
    for farmer_data in farmers_data:
        try:
            # Check if phone already exists
            existing = await db.farmers.find_one({"phone": farmer_data.phone}, {"_id": 0})
            if existing:
                results["failed"] += 1
                results["errors"].append(f"Phone already exists: {farmer_data.phone}")
                continue
            
            farmer_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc).isoformat()
            
            farmer_doc = {
                "id": farmer_id,
                "name": farmer_data.name,
                "phone": farmer_data.phone,
                "address": farmer_data.address or "",
                "village": farmer_data.village or "",
                "bank_account": farmer_data.bank_account or "",
                "ifsc_code": farmer_data.ifsc_code or "",
                "aadhar_number": farmer_data.aadhar_number or "",
                "total_milk": 0.0,
                "total_due": 0.0,
                "total_paid": 0.0,
                "balance": 0.0,
                "created_at": now,
                "is_active": True
            }
            
            await db.farmers.insert_one(farmer_doc)
            results["success"] += 1
            
        except Exception as e:
            results["failed"] += 1
            results["errors"].append(f"Error adding {farmer_data.name}: {str(e)}")
    
    return results

@api_router.get("/bulk/template/collections")
async def get_collection_template():
    """Get CSV template for bulk collection upload"""
    template = "farmer_phone,shift,quantity,fat,snf\n"
    template += "9876543210,morning,5.5,4.2,8.5\n"
    template += "9876543211,evening,6.0,4.5,8.7\n"
    return {"template": template, "columns": ["farmer_phone", "shift", "quantity", "fat", "snf"]}

@api_router.get("/bulk/template/farmers")
async def get_farmer_template():
    """Get CSV template for bulk farmer upload"""
    template = "name,phone,village,address,bank_account,ifsc_code,aadhar_number\n"
    template += "रामलाल,9876543210,गोकुलपुर,मुख्य बाजार,1234567890,SBIN0001234,123456789012\n"
    return {"template": template, "columns": ["name", "phone", "village", "address", "bank_account", "ifsc_code", "aadhar_number"]}

@api_router.post("/bulk/upload-file")
async def bulk_upload_file(
    file: UploadFile = File(...),
    upload_type: str = "collections",
    current_user: dict = Depends(get_current_user)
):
    """Upload Excel/CSV file for bulk data import"""
    import io
    
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    
    content = await file.read()
    rows = []
    
    try:
        if ext == "csv":
            import csv
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(row) for row in reader]
        elif ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = str(val).strip() if val is not None else ""
                if any(row_dict.values()):
                    rows.append(row_dict)
            wb.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing file: {str(e)}")
    
    if not rows:
        raise HTTPException(status_code=400, detail="No data found in file")
    
    results = {"success": 0, "failed": 0, "errors": []}
    
    if upload_type == "collections":
        for row in rows:
            try:
                phone = row.get("farmer_phone", row.get("phone", "")).strip()
                farmer = await db.farmers.find_one({"phone": phone}, {"_id": 0})
                if not farmer:
                    results["failed"] += 1
                    results["errors"].append(f"Farmer not found: {phone}")
                    continue
                
                shift = row.get("shift", "morning").strip().lower()
                quantity = float(row.get("quantity", 0))
                fat = float(row.get("fat", 0))
                snf_val = row.get("snf", "")
                snf = float(snf_val) if snf_val else calculate_snf(fat)
                
                today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                existing = await db.milk_collections.find_one({
                    "farmer_id": farmer["id"], "date": today, "shift": shift
                }, {"_id": 0})
                
                if existing:
                    results["failed"] += 1
                    results["errors"].append(f"Duplicate: {farmer['name']} ({shift})")
                    continue
                
                rate = await get_milk_rate(fat, snf)
                amount = round(quantity * rate, 2)
                collection_id = str(uuid.uuid4())
                now = datetime.now(timezone.utc)
                
                await db.milk_collections.insert_one({
                    "id": collection_id, "farmer_id": farmer["id"],
                    "farmer_name": farmer["name"], "shift": shift,
                    "quantity": quantity, "fat": fat, "snf": snf,
                    "rate": rate, "amount": amount, "date": today,
                    "created_at": now.isoformat()
                })
                
                await db.farmers.update_one(
                    {"id": farmer["id"]},
                    {"$inc": {"total_milk": quantity, "total_due": amount, "balance": amount}}
                )
                results["success"] += 1
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(str(e))
    else:
        for row in rows:
            try:
                name = row.get("name", "").strip()
                phone = row.get("phone", "").strip()
                if not name or not phone:
                    results["failed"] += 1
                    results["errors"].append(f"Missing name/phone in row")
                    continue
                
                existing = await db.farmers.find_one({"phone": phone}, {"_id": 0})
                if existing:
                    results["failed"] += 1
                    results["errors"].append(f"Phone exists: {phone}")
                    continue
                
                farmer_id = str(uuid.uuid4())
                now = datetime.now(timezone.utc).isoformat()
                await db.farmers.insert_one({
                    "id": farmer_id, "name": name, "phone": phone,
                    "address": row.get("address", ""), "village": row.get("village", ""),
                    "bank_account": row.get("bank_account", ""), "ifsc_code": row.get("ifsc_code", ""),
                    "aadhar_number": row.get("aadhar_number", ""),
                    "total_milk": 0.0, "total_due": 0.0, "total_paid": 0.0, "balance": 0.0,
                    "created_at": now, "is_active": True
                })
                results["success"] += 1
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(str(e))
    
    return results

# ==================== WHATSAPP SHARING ROUTES ====================

@api_router.get("/share/farmer-bill/{farmer_id}")
async def get_farmer_bill_share_link(
    farmer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Generate WhatsApp share link for farmer bill"""
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    # Default to current month
    if not start_date:
        today = datetime.now(timezone.utc)
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get collections and calculate totals
    collections = await db.milk_collections.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(1000)
    
    payments = await db.payments.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(1000)
    
    total_milk = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    total_paid = sum(p["amount"] for p in payments)
    balance = total_amount - total_paid
    
    # Generate WhatsApp message
    message = f"""🥛 *Nirbani Dairy - किसान बिल*

*{farmer['name']}*
📞 {farmer['phone']}

📅 अवधि: {start_date} से {end_date}

📊 *विवरण:*
• कुल दूध: {total_milk:.1f} लीटर
• कुल राशि: ₹{total_amount:.2f}
• भुगतान: ₹{total_paid:.2f}
• *बकाया: ₹{balance:.2f}*

धन्यवाद! 🙏"""

    # URL encode the message
    import urllib.parse
    encoded_message = urllib.parse.quote(message)
    whatsapp_link = f"https://wa.me/{farmer['phone']}?text={encoded_message}"
    
    return {
        "farmer": farmer["name"],
        "phone": farmer["phone"],
        "message": message,
        "whatsapp_link": whatsapp_link,
        "summary": {
            "total_milk": round(total_milk, 2),
            "total_amount": round(total_amount, 2),
            "total_paid": round(total_paid, 2),
            "balance": round(balance, 2)
        }
    }

@api_router.get("/share/daily-report/{date}")
async def get_daily_report_share(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Generate WhatsApp share link for daily report"""
    collections = await db.milk_collections.find({"date": date}, {"_id": 0}).to_list(1000)
    
    total_quantity = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    morning_qty = sum(c["quantity"] for c in collections if c["shift"] == "morning")
    evening_qty = sum(c["quantity"] for c in collections if c["shift"] == "evening")
    unique_farmers = len(set(c["farmer_id"] for c in collections))
    
    if collections:
        avg_fat = sum(c["fat"] for c in collections) / len(collections)
    else:
        avg_fat = 0
    
    message = f"""🥛 *Nirbani Dairy - दैनिक रिपोर्ट*

📅 तारीख: {date}

📊 *संग्रह विवरण:*
• कुल दूध: {total_quantity:.1f} लीटर
• सुबह: {morning_qty:.1f} L | शाम: {evening_qty:.1f} L
• औसत फैट: {avg_fat:.1f}%
• कुल राशि: ₹{total_amount:.2f}

👥 किसान: {unique_farmers}
📝 प्रविष्टियाँ: {len(collections)}"""

    import urllib.parse
    encoded_message = urllib.parse.quote(message)
    
    return {
        "date": date,
        "message": message,
        "whatsapp_link": f"https://wa.me/?text={encoded_message}",
        "summary": {
            "total_quantity": round(total_quantity, 2),
            "total_amount": round(total_amount, 2),
            "morning_quantity": round(morning_qty, 2),
            "evening_quantity": round(evening_qty, 2),
            "avg_fat": round(avg_fat, 2),
            "farmers": unique_farmers,
            "entries": len(collections)
        }
    }

# ==================== DASHBOARD ROUTES ====================

@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Count farmers
    total_farmers = await db.farmers.count_documents({})
    active_farmers = await db.farmers.count_documents({"is_active": True})
    
    # Today's collections
    today_collections = await db.milk_collections.find({"date": today}, {"_id": 0}).to_list(1000)
    
    today_milk_quantity = sum(c["quantity"] for c in today_collections)
    today_milk_amount = sum(c["amount"] for c in today_collections)
    today_morning = sum(c["quantity"] for c in today_collections if c["shift"] == "morning")
    today_evening = sum(c["quantity"] for c in today_collections if c["shift"] == "evening")
    
    # Calculate averages
    if today_collections:
        avg_fat = sum(c["fat"] for c in today_collections) / len(today_collections)
        avg_snf = sum(c["snf"] for c in today_collections) / len(today_collections)
    else:
        avg_fat = 0
        avg_snf = 0
    
    # Pending payments (total balance across all farmers)
    pipeline = [{"$group": {"_id": None, "total": {"$sum": "$balance"}}}]
    balance_result = await db.farmers.aggregate(pipeline).to_list(1)
    total_pending = balance_result[0]["total"] if balance_result else 0
    
    return DashboardStats(
        total_farmers=total_farmers,
        active_farmers=active_farmers,
        today_milk_quantity=round(today_milk_quantity, 2),
        today_milk_amount=round(today_milk_amount, 2),
        today_morning_quantity=round(today_morning, 2),
        today_evening_quantity=round(today_evening, 2),
        avg_fat=round(avg_fat, 2),
        avg_snf=round(avg_snf, 2),
        total_pending_payments=round(total_pending, 2),
        collections_count=len(today_collections)
    )

@api_router.get("/dashboard/weekly-stats")
async def get_weekly_stats(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc)
    week_ago = today - timedelta(days=7)
    
    dates = []
    for i in range(7):
        d = week_ago + timedelta(days=i+1)
        dates.append(d.strftime("%Y-%m-%d"))
    
    stats = []
    for date in dates:
        collections = await db.milk_collections.find({"date": date}, {"_id": 0}).to_list(1000)
        quantity = sum(c["quantity"] for c in collections)
        amount = sum(c["amount"] for c in collections)
        stats.append({
            "date": date,
            "quantity": round(quantity, 2),
            "amount": round(amount, 2),
            "count": len(collections)
        })
    
    return stats

# ==================== REPORTS ROUTES ====================

@api_router.get("/reports/daily")
async def get_daily_report(
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    collections = await db.milk_collections.find({"date": date}, {"_id": 0}).to_list(1000)
    payments = await db.payments.find({"date": date}, {"_id": 0}).to_list(1000)
    
    total_quantity = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    morning_qty = sum(c["quantity"] for c in collections if c["shift"] == "morning")
    evening_qty = sum(c["quantity"] for c in collections if c["shift"] == "evening")
    total_paid = sum(p["amount"] for p in payments)
    
    return {
        "date": date,
        "collections": collections,
        "payments": payments,
        "summary": {
            "total_quantity": round(total_quantity, 2),
            "total_amount": round(total_amount, 2),
            "morning_quantity": round(morning_qty, 2),
            "evening_quantity": round(evening_qty, 2),
            "collection_count": len(collections),
            "total_paid": round(total_paid, 2),
            "payment_count": len(payments)
        }
    }

@api_router.get("/reports/farmer/{farmer_id}")
async def get_farmer_report(
    farmer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    query = {"farmer_id": farmer_id}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    collections = await db.milk_collections.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    payment_query = {"farmer_id": farmer_id}
    if start_date:
        payment_query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in payment_query:
            payment_query["date"]["$lte"] = end_date
        else:
            payment_query["date"] = {"$lte": end_date}
    
    payments = await db.payments.find(payment_query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    period_milk = sum(c["quantity"] for c in collections)
    period_amount = sum(c["amount"] for c in collections)
    period_paid = sum(p["amount"] for p in payments)
    
    return {
        "farmer": FarmerResponse(**farmer),
        "period": {"start": start_date, "end": end_date},
        "collections": collections,
        "payments": payments,
        "summary": {
            "total_milk": round(period_milk, 2),
            "total_amount": round(period_amount, 2),
            "total_paid": round(period_paid, 2),
            "balance": round(period_amount - period_paid, 2),
            "collection_count": len(collections),
            "payment_count": len(payments)
        }
    }

# ==================== ADVANCED REPORTS ====================

@api_router.get("/reports/fat-average")
async def get_fat_average_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get fat average report by farmer"""
    if not start_date:
        today = datetime.now(timezone.utc)
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    collections = await db.milk_collections.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(10000)
    
    # Group by farmer
    farmer_data = {}
    for c in collections:
        fid = c["farmer_id"]
        if fid not in farmer_data:
            farmer_data[fid] = {
                "farmer_name": c["farmer_name"],
                "total_quantity": 0,
                "total_fat": 0,
                "total_snf": 0,
                "count": 0
            }
        farmer_data[fid]["total_quantity"] += c["quantity"]
        farmer_data[fid]["total_fat"] += c["fat"] * c["quantity"]
        farmer_data[fid]["total_snf"] += c["snf"] * c["quantity"]
        farmer_data[fid]["count"] += 1
    
    # Calculate weighted averages
    report = []
    for fid, data in farmer_data.items():
        if data["total_quantity"] > 0:
            avg_fat = data["total_fat"] / data["total_quantity"]
            avg_snf = data["total_snf"] / data["total_quantity"]
        else:
            avg_fat = 0
            avg_snf = 0
        
        report.append({
            "farmer_id": fid,
            "farmer_name": data["farmer_name"],
            "total_quantity": round(data["total_quantity"], 2),
            "avg_fat": round(avg_fat, 2),
            "avg_snf": round(avg_snf, 2),
            "entries": data["count"]
        })
    
    # Sort by average fat descending
    report.sort(key=lambda x: x["avg_fat"], reverse=True)
    
    return {
        "period": {"start": start_date, "end": end_date},
        "report": report
    }

@api_router.get("/reports/farmer-ranking")
async def get_farmer_ranking_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    sort_by: str = "quantity",  # quantity, amount, fat
    current_user: dict = Depends(get_current_user)
):
    """Get farmer ranking report"""
    if not start_date:
        today = datetime.now(timezone.utc)
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    collections = await db.milk_collections.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(10000)
    
    # Group by farmer
    farmer_data = {}
    for c in collections:
        fid = c["farmer_id"]
        if fid not in farmer_data:
            farmer_data[fid] = {
                "farmer_name": c["farmer_name"],
                "total_quantity": 0,
                "total_amount": 0,
                "total_fat_weighted": 0,
                "count": 0
            }
        farmer_data[fid]["total_quantity"] += c["quantity"]
        farmer_data[fid]["total_amount"] += c["amount"]
        farmer_data[fid]["total_fat_weighted"] += c["fat"] * c["quantity"]
        farmer_data[fid]["count"] += 1
    
    # Build ranking list
    ranking = []
    for fid, data in farmer_data.items():
        avg_fat = data["total_fat_weighted"] / data["total_quantity"] if data["total_quantity"] > 0 else 0
        ranking.append({
            "farmer_id": fid,
            "farmer_name": data["farmer_name"],
            "total_quantity": round(data["total_quantity"], 2),
            "total_amount": round(data["total_amount"], 2),
            "avg_fat": round(avg_fat, 2),
            "entries": data["count"]
        })
    
    # Sort based on criteria
    if sort_by == "quantity":
        ranking.sort(key=lambda x: x["total_quantity"], reverse=True)
    elif sort_by == "amount":
        ranking.sort(key=lambda x: x["total_amount"], reverse=True)
    elif sort_by == "fat":
        ranking.sort(key=lambda x: x["avg_fat"], reverse=True)
    
    # Add rank
    for i, r in enumerate(ranking):
        r["rank"] = i + 1
    
    return {
        "period": {"start": start_date, "end": end_date},
        "sort_by": sort_by,
        "ranking": ranking
    }

@api_router.get("/reports/monthly-summary")
async def get_monthly_summary_report(
    month: Optional[str] = None,  # Format: YYYY-MM
    current_user: dict = Depends(get_current_user)
):
    """Get monthly business summary"""
    if not month:
        month = datetime.now(timezone.utc).strftime("%Y-%m")
    
    year, mon = month.split("-")
    start_date = f"{month}-01"
    
    # Calculate end date (last day of month)
    if int(mon) == 12:
        end_date = f"{int(year)+1}-01-01"
    else:
        end_date = f"{year}-{int(mon)+1:02d}-01"
    
    # Get all data for the month
    collections = await db.milk_collections.find(
        {"date": {"$gte": start_date, "$lt": end_date}}, {"_id": 0}
    ).to_list(10000)
    
    payments = await db.payments.find(
        {"date": {"$gte": start_date, "$lt": end_date}}, {"_id": 0}
    ).to_list(10000)
    
    sales = await db.sales.find(
        {"date": {"$gte": start_date, "$lt": end_date}}, {"_id": 0}
    ).to_list(10000)
    
    expenses = await db.expenses.find(
        {"date": {"$gte": start_date, "$lt": end_date}}, {"_id": 0}
    ).to_list(10000)
    
    # Calculate totals
    total_milk = sum(c["quantity"] for c in collections)
    total_milk_amount = sum(c["amount"] for c in collections)
    total_payments = sum(p["amount"] for p in payments)
    total_sales = sum(s["amount"] for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    
    # Unique farmers
    unique_farmers = len(set(c["farmer_id"] for c in collections))
    
    # Average fat
    if collections:
        avg_fat = sum(c["fat"] * c["quantity"] for c in collections) / total_milk if total_milk > 0 else 0
    else:
        avg_fat = 0
    
    # Daily breakdown
    daily_data = {}
    for c in collections:
        if c["date"] not in daily_data:
            daily_data[c["date"]] = {"quantity": 0, "amount": 0}
        daily_data[c["date"]]["quantity"] += c["quantity"]
        daily_data[c["date"]]["amount"] += c["amount"]
    
    return {
        "month": month,
        "collection": {
            "total_milk": round(total_milk, 2),
            "total_amount": round(total_milk_amount, 2),
            "unique_farmers": unique_farmers,
            "avg_fat": round(avg_fat, 2),
            "entries": len(collections)
        },
        "payments": {
            "total": round(total_payments, 2),
            "count": len(payments)
        },
        "sales": {
            "total": round(total_sales, 2),
            "count": len(sales)
        },
        "expenses": {
            "total": round(total_expenses, 2),
            "count": len(expenses)
        },
        "profit": {
            "gross": round(total_sales - total_milk_amount, 2),
            "net": round(total_sales - total_milk_amount - total_expenses, 2)
        },
        "daily_breakdown": [
            {"date": d, **v} for d, v in sorted(daily_data.items())
        ]
    }

# ==================== BILL GENERATION ROUTES ====================

@api_router.get("/bills/farmer/{farmer_id}", response_class=HTMLResponse)
async def generate_farmer_bill(
    farmer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Generate HTML bill for a farmer"""
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    # Default to current month if no dates
    if not start_date:
        today = datetime.now(timezone.utc)
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get collections
    collection_query = {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}
    collections = await db.milk_collections.find(collection_query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    # Get payments
    payment_query = {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}
    payments = await db.payments.find(payment_query, {"_id": 0}).sort("date", 1).to_list(1000)
    
    # Get settings for dairy info
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    
    html = generate_farmer_bill_html(
        farmer=farmer,
        collections=collections,
        payments=payments,
        period_start=start_date,
        period_end=end_date,
        dairy_name=settings.get("dairy_name", "Nirbani Dairy"),
        dairy_phone=settings.get("dairy_phone", ""),
        dairy_address=settings.get("dairy_address", "")
    )
    
    return HTMLResponse(content=html)

@api_router.get("/bills/daily/{date}", response_class=HTMLResponse)
async def generate_daily_bill(
    date: str,
    current_user: dict = Depends(get_current_user)
):
    """Generate HTML daily report"""
    collections = await db.milk_collections.find({"date": date}, {"_id": 0}).to_list(1000)
    payments = await db.payments.find({"date": date}, {"_id": 0}).to_list(1000)
    
    total_quantity = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    morning_qty = sum(c["quantity"] for c in collections if c["shift"] == "morning")
    evening_qty = sum(c["quantity"] for c in collections if c["shift"] == "evening")
    
    summary = {
        "total_quantity": total_quantity,
        "total_amount": total_amount,
        "morning_quantity": morning_qty,
        "evening_quantity": evening_qty
    }
    
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    
    html = generate_daily_report_html(
        date=date,
        collections=collections,
        payments=payments,
        summary=summary,
        dairy_name=settings.get("dairy_name", "Nirbani Dairy")
    )
    
    return HTMLResponse(content=html)

# ==================== SETTINGS ROUTES ====================

class DairySettings(BaseModel):
    dairy_name: str = "Nirbani Dairy"
    dairy_phone: str = ""
    dairy_address: str = ""
    sms_enabled: bool = False
    msg91_auth_key: str = ""
    msg91_sender_id: str = "NIRDRY"

class SMSTemplateSettings(BaseModel):
    collection_template: str = ""
    payment_template: str = ""

@api_router.get("/settings/dairy")
async def get_dairy_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0})
    if not settings:
        return {
            "dairy_name": "Nirbani Dairy",
            "dairy_phone": "",
            "dairy_address": "",
            "sms_enabled": False
        }
    return settings

@api_router.put("/settings/dairy")
async def update_dairy_settings(
    settings: DairySettings,
    current_user: dict = Depends(get_current_user)
):
    await db.settings.update_one(
        {"type": "dairy_info"},
        {"$set": {
            "type": "dairy_info",
            "dairy_name": settings.dairy_name,
            "dairy_phone": settings.dairy_phone,
            "dairy_address": settings.dairy_address,
            "sms_enabled": settings.sms_enabled,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    return {"message": "Settings updated successfully"}

@api_router.get("/settings/sms-templates")
async def get_sms_templates(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find_one({"type": "sms_templates"}, {"_id": 0})
    if not settings:
        return {
            "collection_template": "Nirbani Dairy: {farmer_name} जी, आपका {shift} का दूध: मात्रा: {quantity}L | फैट: {fat}% | राशि: ₹{amount}",
            "payment_template": "Nirbani Dairy: {farmer_name} जी, ₹{amount} का भुगतान प्राप्त। बकाया: ₹{balance}"
        }
    return settings

@api_router.put("/settings/sms-templates")
async def update_sms_templates(
    templates: SMSTemplateSettings,
    current_user: dict = Depends(get_current_user)
):
    await db.settings.update_one(
        {"type": "sms_templates"},
        {"$set": {
            "type": "sms_templates",
            "collection_template": templates.collection_template,
            "payment_template": templates.payment_template,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    return {"message": "SMS templates updated successfully"}

# ==================== EXPORT ROUTES ====================

@api_router.get("/export/collections")
async def export_collections(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export collections as CSV"""
    import io, csv
    if not start_date:
        start_date = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    collections = await db.milk_collections.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Farmer", "Shift", "Quantity(L)", "Fat%", "SNF%", "Rate", "Amount"])
    for c in collections:
        writer.writerow([c["date"], c["farmer_name"], c["shift"], c["quantity"], c["fat"], c["snf"], c["rate"], c["amount"]])
    
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=collections_{start_date}_to_{end_date}.csv"}
    )

@api_router.get("/export/payments")
async def export_payments(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export payments as CSV"""
    import io, csv
    if not start_date:
        start_date = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    payments = await db.payments.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Farmer", "Amount", "Mode", "Type", "Notes"])
    for p in payments:
        writer.writerow([p["date"], p["farmer_name"], p["amount"], p["payment_mode"], p.get("payment_type", "payment"), p.get("notes", "")])
    
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=payments_{start_date}_to_{end_date}.csv"}
    )

@api_router.get("/export/farmers")
async def export_farmers(current_user: dict = Depends(get_current_user)):
    """Export farmers list as CSV"""
    import io, csv
    farmers = await db.farmers.find({}, {"_id": 0}).sort("name", 1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Name", "Phone", "Village", "Address", "Total Milk(L)", "Total Due", "Total Paid", "Balance", "Active"])
    for f in farmers:
        writer.writerow([f["name"], f["phone"], f.get("village",""), f.get("address",""),
            f["total_milk"], f["total_due"], f["total_paid"], f["balance"], f.get("is_active", True)])
    
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=farmers_list.csv"}
    )

@api_router.get("/export/sales")
async def export_sales(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export sales as CSV"""
    import io, csv
    if not start_date:
        start_date = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    sales = await db.sales.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Customer", "Product", "Quantity", "Rate", "Amount"])
    for s in sales:
        writer.writerow([s["date"], s["customer_name"], s["product"], s["quantity"], s["rate"], s["amount"]])
    
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=sales_{start_date}_to_{end_date}.csv"}
    )

@api_router.get("/export/expenses")
async def export_expenses(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export expenses as CSV"""
    import io, csv
    if not start_date:
        start_date = datetime.now(timezone.utc).replace(day=1).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    expenses = await db.expenses.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(10000)
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date", "Category", "Amount", "Description", "Payment Mode"])
    for e in expenses:
        writer.writerow([e["date"], e["category"], e["amount"], e.get("description",""), e["payment_mode"]])
    
    content = output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=expenses_{start_date}_to_{end_date}.csv"}
    )

# ==================== THERMAL PRINTER BILL ====================

@api_router.get("/bills/thermal/{farmer_id}", response_class=HTMLResponse)
async def thermal_bill(farmer_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Generate thermal printer friendly bill (58mm/80mm)"""
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    collections = await db.milk_collections.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    payments = await db.payments.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    dairy_phone = settings.get("phone", "")
    
    total_milk = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    total_paid = sum(p["amount"] for p in payments)
    balance = total_amount - total_paid
    
    rows = ""
    for c in collections:
        shift = "AM" if c["shift"] == "morning" else "PM"
        rows += f"<tr><td>{c['date'][5:]}</td><td>{shift}</td><td>{c['quantity']}</td><td>{c['fat']}</td><td>{c['rate']}</td><td style='text-align:right'>{c['amount']:.0f}</td></tr>\n"
    
    pay_rows = ""
    for p in payments:
        pay_rows += f"<tr><td>{p['date'][5:]}</td><td>{p['payment_mode']}</td><td style='text-align:right'>{p['amount']:.0f}</td></tr>\n"
    
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><meta name="viewport" content="width=58mm">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Courier New',monospace;font-size:11px;width:58mm;padding:2mm;color:#000}}
.center{{text-align:center}}.bold{{font-weight:bold}}.line{{border-top:1px dashed #000;margin:3px 0}}
table{{width:100%;border-collapse:collapse}}
td,th{{padding:1px 2px;font-size:10px;vertical-align:top}}
th{{text-align:left;border-bottom:1px solid #000}}
.right{{text-align:right}}.big{{font-size:14px}}
@media print{{@page{{size:58mm auto;margin:0}}body{{width:58mm}}}}
</style></head><body>
<div class="center bold big">{dairy_name}</div>
<div class="center" style="font-size:9px">{dairy_phone}</div>
<div class="line"></div>
<div class="bold">{farmer['name']}</div>
<div style="font-size:9px">Ph: {farmer['phone']}</div>
<div style="font-size:9px">{start_date} to {end_date}</div>
<div class="line"></div>
<table><tr><th>Date</th><th>S</th><th>Qty</th><th>Fat</th><th>Rate</th><th class="right">Amt</th></tr>
{rows}</table>
<div class="line"></div>
<table>
<tr><td class="bold">Total Milk:</td><td class="right bold">{total_milk:.1f} L</td></tr>
<tr><td class="bold">Total Amt:</td><td class="right bold">Rs.{total_amount:.0f}</td></tr>
<tr><td class="bold">Paid:</td><td class="right bold">Rs.{total_paid:.0f}</td></tr>
<tr><td class="bold">Balance:</td><td class="right bold">Rs.{balance:.0f}</td></tr>
</table>
{f'<div class="line"></div><div class="bold" style="font-size:9px">Payments:</div><table><tr><th>Date</th><th>Mode</th><th class="right">Amt</th></tr>{pay_rows}</table>' if payments else ''}
<div class="line"></div>
<div class="center" style="font-size:9px;margin-top:3px">Thank You / धन्यवाद</div>
<div class="center" style="font-size:8px">Printed: {datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M")}</div>
</body></html>"""
    return HTMLResponse(content=html)

# ==================== A4 INVOICE ====================

@api_router.get("/bills/a4/{farmer_id}", response_class=HTMLResponse)
async def a4_invoice(farmer_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """Generate A4 professional invoice"""
    farmer = await db.farmers.find_one({"id": farmer_id}, {"_id": 0})
    if not farmer:
        raise HTTPException(status_code=404, detail="Farmer not found")
    
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    
    collections = await db.milk_collections.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    payments = await db.payments.find(
        {"farmer_id": farmer_id, "date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).sort("date", 1).to_list(1000)
    
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    dairy_phone = settings.get("phone", "")
    dairy_address = settings.get("address", "")
    
    total_milk = sum(c["quantity"] for c in collections)
    total_amount = sum(c["amount"] for c in collections)
    avg_fat = sum(c["fat"] * c["quantity"] for c in collections) / total_milk if total_milk > 0 else 0
    avg_snf = sum(c["snf"] * c["quantity"] for c in collections) / total_milk if total_milk > 0 else 0
    total_paid = sum(p["amount"] for p in payments)
    balance = total_amount - total_paid
    invoice_no = f"INV-{farmer_id[:8].upper()}-{datetime.now(timezone.utc).strftime('%Y%m%d')}"
    
    coll_rows = ""
    for i, c in enumerate(collections, 1):
        shift_label = "Morning / सुबह" if c["shift"] == "morning" else "Evening / शाम"
        coll_rows += f"""<tr><td>{i}</td><td>{c['date']}</td><td>{shift_label}</td><td>{c['quantity']:.1f}</td><td>{c['fat']:.1f}</td><td>{c['snf']:.1f}</td><td>{c['rate']:.2f}</td><td class="right">{c['amount']:.2f}</td></tr>"""
    
    pay_rows = ""
    for p in payments:
        mode = p.get("payment_mode", "cash").upper()
        pay_rows += f"<tr><td>{p['date']}</td><td>{mode}</td><td>{p.get('notes','')}</td><td class='right'>{p['amount']:.2f}</td></tr>"
    
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1a1a1a;padding:15mm 20mm;max-width:210mm}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #15803d;padding-bottom:12px;margin-bottom:15px}}
.dairy-name{{font-size:24px;font-weight:bold;color:#15803d}}.dairy-info{{font-size:10px;color:#666;margin-top:4px}}
.invoice-box{{text-align:right}}.invoice-title{{font-size:20px;color:#15803d;font-weight:bold}}.invoice-no{{font-size:11px;color:#666;margin-top:2px}}
.info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:15px;margin-bottom:15px}}
.info-card{{background:#f8faf8;border:1px solid #e5e7eb;border-radius:6px;padding:10px}}
.info-label{{font-size:9px;color:#888;text-transform:uppercase;letter-spacing:0.5px}}.info-value{{font-size:13px;font-weight:600;margin-top:2px}}
table{{width:100%;border-collapse:collapse;margin-bottom:15px}}
th{{background:#15803d;color:white;padding:6px 8px;text-align:left;font-size:10px;text-transform:uppercase}}
td{{padding:5px 8px;border-bottom:1px solid #eee;font-size:11px}}.right{{text-align:right}}
tr:nth-child(even){{background:#f9fafb}}
.summary-grid{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;margin:15px 0}}
.summary-box{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:10px;text-align:center}}
.summary-box.due{{background:#fef2f2;border-color:#fecaca}}
.summary-label{{font-size:9px;color:#666;text-transform:uppercase}}.summary-value{{font-size:18px;font-weight:bold;color:#15803d;margin-top:2px}}
.summary-box.due .summary-value{{color:#dc2626}}
.footer{{margin-top:20px;border-top:2px solid #15803d;padding-top:10px;display:flex;justify-content:space-between;font-size:10px;color:#666}}
.sig-box{{border-top:1px solid #333;width:150px;text-align:center;padding-top:5px;margin-top:40px;font-size:10px}}
h3{{color:#15803d;font-size:13px;margin-bottom:8px;padding-bottom:4px;border-bottom:1px solid #d1fae5}}
@media print{{@page{{size:A4;margin:10mm}}body{{padding:0}}}}
</style></head><body>
<div class="header">
  <div><div class="dairy-name">{dairy_name}</div><div class="dairy-info">{dairy_address}<br>{dairy_phone}</div></div>
  <div class="invoice-box"><div class="invoice-title">INVOICE / बिल</div><div class="invoice-no">{invoice_no}</div><div class="invoice-no">{start_date} to {end_date}</div></div>
</div>
<div class="info-grid">
  <div class="info-card"><div class="info-label">Farmer / किसान</div><div class="info-value">{farmer['name']}</div><div style="font-size:10px;color:#666">Ph: {farmer['phone']}{f" | Village: {farmer.get('village','')}" if farmer.get('village') else ''}</div></div>
  <div class="info-card"><div class="info-label">Account / Bank</div><div class="info-value">{farmer.get('bank_account','N/A')}</div><div style="font-size:10px;color:#666">IFSC: {farmer.get('ifsc_code','N/A')}</div></div>
</div>
<div class="summary-grid">
  <div class="summary-box"><div class="summary-label">Total Milk / कुल दूध</div><div class="summary-value">{total_milk:.1f} L</div></div>
  <div class="summary-box"><div class="summary-label">Avg Fat / औसत फैट</div><div class="summary-value">{avg_fat:.1f}%</div></div>
  <div class="summary-box"><div class="summary-label">Total Amount / कुल राशि</div><div class="summary-value">₹{total_amount:.0f}</div></div>
  <div class="summary-box {'due' if balance > 0 else ''}"><div class="summary-label">Balance / बकाया</div><div class="summary-value">₹{balance:.0f}</div></div>
</div>
<h3>Milk Collection Details / दूध संग्रह विवरण</h3>
<table><tr><th>#</th><th>Date</th><th>Shift / पाली</th><th>Qty (L)</th><th>Fat %</th><th>SNF %</th><th>Rate ₹/L</th><th class="right">Amount ₹</th></tr>
{coll_rows}
<tr style="background:#15803d;color:white;font-weight:bold"><td colspan="3">TOTAL</td><td>{total_milk:.1f}</td><td>{avg_fat:.1f}</td><td>{avg_snf:.1f}</td><td></td><td class="right">₹{total_amount:.2f}</td></tr>
</table>
{f'<h3>Payments / भुगतान</h3><table><tr><th>Date</th><th>Mode</th><th>Notes</th><th class="right">Amount ₹</th></tr>{pay_rows}<tr style="background:#15803d;color:white;font-weight:bold"><td colspan="3">TOTAL PAID</td><td class="right">₹{total_paid:.2f}</td></tr></table>' if payments else ''}
<div style="display:flex;justify-content:flex-end;gap:40px;margin-top:30px">
  <div class="sig-box">Dairy Stamp / डेयरी मुहर</div>
  <div class="sig-box">Authorized Signature / हस्ताक्षर</div>
</div>
<div class="footer"><div>Generated: {datetime.now(timezone.utc).strftime("%d-%m-%Y %H:%M UTC")}</div><div>{dairy_name} | {dairy_phone}</div></div>
</body></html>"""
    return HTMLResponse(content=html)

# ==================== OCR RATE CHART UPLOAD ====================

@api_router.post("/rate-charts/ocr-upload")
async def ocr_rate_chart_upload(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload image/PDF of rate chart and extract fat/SNF rates using AI OCR"""
    import base64
    
    llm_key = os.environ.get("EMERGENT_LLM_KEY")
    if not llm_key:
        raise HTTPException(status_code=500, detail="LLM key not configured")
    
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("jpg", "jpeg", "png", "webp", "pdf"):
        raise HTTPException(status_code=400, detail="Only JPG, PNG, WEBP, or PDF files are supported")
    
    content = await file.read()
    
    if ext == "pdf":
        raise HTTPException(status_code=400, detail="Please upload an image (JPG/PNG/WEBP) of the rate chart. PDF OCR coming soon.")
    
    image_b64 = base64.b64encode(content).decode("utf-8")
    
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
        
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"ocr-{uuid.uuid4()}",
            system_message="""You are a dairy rate chart OCR expert. Extract milk rate data from images.
Return ONLY a valid JSON array where each entry has: {"fat": number, "snf": number, "rate": number}
Example: [{"fat": 3.0, "snf": 8.0, "rate": 25.50}, {"fat": 3.5, "snf": 8.5, "rate": 28.00}]
If SNF is not visible, estimate it as (fat * 2) + 0.5.
Return ONLY the JSON array, no markdown, no explanation."""
        )
        chat.with_model("openai", "gpt-4o")
        
        image_content = ImageContent(image_base64=image_b64)
        user_msg = UserMessage(
            text="Extract ALL fat, SNF, and rate values from this dairy milk rate chart image. Return as JSON array.",
            file_contents=[image_content]
        )
        
        response = await chat.send_message(user_msg)
        
        import json
        response_text = response.strip()
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
            response_text = response_text.rsplit("```", 1)[0]
        
        rate_data = json.loads(response_text.strip())
        
        if not isinstance(rate_data, list):
            raise ValueError("Response is not a list")
        
        inserted = 0
        for entry in rate_data:
            fat = float(entry.get("fat", 0))
            snf = float(entry.get("snf", 0))
            rate = float(entry.get("rate", 0))
            if fat > 0 and rate > 0:
                existing = await db.rate_charts.find_one({"fat": fat, "snf": snf}, {"_id": 0})
                if existing:
                    await db.rate_charts.update_one({"fat": fat, "snf": snf}, {"$set": {"rate": rate}})
                else:
                    await db.rate_charts.insert_one({
                        "id": str(uuid.uuid4()), "fat": fat, "snf": snf,
                        "rate": rate, "created_at": datetime.now(timezone.utc).isoformat()
                    })
                inserted += 1
        
        return {
            "success": True, "extracted": len(rate_data), "saved": inserted,
            "rates": rate_data,
            "message": f"Successfully extracted {len(rate_data)} rates, saved {inserted} to rate chart"
        }
    
    except json.JSONDecodeError:
        return {"success": False, "error": "Could not parse rate data from image. Please try with a clearer image."}
    except Exception as e:
        logging.error(f"OCR Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

# ==================== DAIRY PLANT ROUTES ====================

@api_router.post("/dairy-plants", response_model=DairyPlantResponse)
async def create_dairy_plant(plant: DairyPlantCreate, current_user: dict = Depends(get_current_user)):
    plant_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    plant_doc = {
        "id": plant_id,
        "name": plant.name,
        "code": plant.code or "",
        "address": plant.address or "",
        "phone": plant.phone or "",
        "contact_person": plant.contact_person or "",
        "total_milk_supplied": 0,
        "total_amount": 0,
        "total_paid": 0,
        "balance": 0,
        "is_active": True,
        "created_at": now.isoformat(),
    }
    await db.dairy_plants.insert_one(plant_doc)
    del plant_doc["_id"]
    return DairyPlantResponse(**plant_doc)

@api_router.get("/dairy-plants", response_model=List[DairyPlantResponse])
async def get_dairy_plants(current_user: dict = Depends(get_current_user)):
    plants = await db.dairy_plants.find({"is_active": True}, {"_id": 0}).sort("name", 1).to_list(100)
    return [DairyPlantResponse(**p) for p in plants]

@api_router.get("/dairy-plants/{plant_id}", response_model=DairyPlantResponse)
async def get_dairy_plant(plant_id: str, current_user: dict = Depends(get_current_user)):
    plant = await db.dairy_plants.find_one({"id": plant_id}, {"_id": 0})
    if not plant:
        raise HTTPException(status_code=404, detail="Dairy plant not found")
    return DairyPlantResponse(**plant)

@api_router.put("/dairy-plants/{plant_id}", response_model=DairyPlantResponse)
async def update_dairy_plant(plant_id: str, updates: DairyPlantUpdate, current_user: dict = Depends(get_current_user)):
    plant = await db.dairy_plants.find_one({"id": plant_id}, {"_id": 0})
    if not plant:
        raise HTTPException(status_code=404, detail="Dairy plant not found")
    update_data = {k: v for k, v in updates.model_dump().items() if v is not None}
    if update_data:
        await db.dairy_plants.update_one({"id": plant_id}, {"$set": update_data})
    updated = await db.dairy_plants.find_one({"id": plant_id}, {"_id": 0})
    return DairyPlantResponse(**updated)

# ==================== DISPATCH ROUTES ====================

@api_router.post("/dispatches", response_model=DispatchResponse)
async def create_dispatch(dispatch: DispatchCreate, current_user: dict = Depends(get_current_user)):
    plant = await db.dairy_plants.find_one({"id": dispatch.dairy_plant_id}, {"_id": 0})
    if not plant:
        raise HTTPException(status_code=404, detail="Dairy plant not found")

    dispatch_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    date_str = dispatch.date or now.strftime("%Y-%m-%d")

    gross_amount = round(dispatch.quantity_kg * dispatch.rate_per_kg, 2)
    deductions_list = [d.model_dump() for d in (dispatch.deductions or [])]
    total_deduction = round(sum(d["amount"] for d in deductions_list), 2)
    net_receivable = round(gross_amount - total_deduction, 2)

    dispatch_doc = {
        "id": dispatch_id,
        "dairy_plant_id": dispatch.dairy_plant_id,
        "dairy_plant_name": plant["name"],
        "date": date_str,
        "tanker_number": dispatch.tanker_number or "",
        "quantity_kg": dispatch.quantity_kg,
        "avg_fat": dispatch.avg_fat,
        "avg_snf": dispatch.avg_snf,
        "clr": dispatch.clr,
        "rate_per_kg": dispatch.rate_per_kg,
        "gross_amount": gross_amount,
        "deductions": deductions_list,
        "total_deduction": total_deduction,
        "net_receivable": net_receivable,
        "notes": dispatch.notes or "",
        "slip_fat": None,
        "slip_snf": None,
        "slip_amount": None,
        "slip_deductions": None,
        "fat_difference": None,
        "amount_difference": None,
        "slip_matched": False,
        "created_at": now.isoformat(),
    }
    await db.dispatches.insert_one(dispatch_doc)
    del dispatch_doc["_id"]

    # Update dairy plant totals
    await db.dairy_plants.update_one(
        {"id": dispatch.dairy_plant_id},
        {"$inc": {"total_milk_supplied": dispatch.quantity_kg, "total_amount": net_receivable, "balance": net_receivable}}
    )

    return DispatchResponse(**dispatch_doc)

@api_router.get("/dispatches")
async def get_dispatches(
    dairy_plant_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if dairy_plant_id:
        query["dairy_plant_id"] = dairy_plant_id
    if start_date:
        query.setdefault("date", {})["$gte"] = start_date
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date
    dispatches = await db.dispatches.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return dispatches

@api_router.get("/dispatches/{dispatch_id}", response_model=DispatchResponse)
async def get_dispatch(dispatch_id: str, current_user: dict = Depends(get_current_user)):
    dispatch = await db.dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    return DispatchResponse(**dispatch)

@api_router.delete("/dispatches/{dispatch_id}")
async def delete_dispatch(dispatch_id: str, current_user: dict = Depends(get_current_user)):
    dispatch = await db.dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    await db.dairy_plants.update_one(
        {"id": dispatch["dairy_plant_id"]},
        {"$inc": {"total_milk_supplied": -dispatch["quantity_kg"], "total_amount": -dispatch["net_receivable"], "balance": -dispatch["net_receivable"]}}
    )
    await db.dispatches.delete_one({"id": dispatch_id})
    return {"message": "Dispatch deleted"}

# ==================== SLIP MATCHING ROUTES ====================

@api_router.put("/dispatches/{dispatch_id}/slip-match")
async def match_dispatch_slip(dispatch_id: str, slip: SlipMatchCreate, current_user: dict = Depends(get_current_user)):
    dispatch = await db.dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")

    fat_diff = round(dispatch["avg_fat"] - slip.slip_fat, 2)
    amount_diff = round(dispatch["net_receivable"] - slip.slip_amount, 2)

    # If slip amount differs, adjust dairy plant balance
    if dispatch.get("slip_matched") and dispatch.get("slip_amount"):
        old_diff = dispatch["net_receivable"] - dispatch["slip_amount"]
        await db.dairy_plants.update_one(
            {"id": dispatch["dairy_plant_id"]},
            {"$inc": {"total_amount": old_diff, "balance": old_diff}}
        )

    new_balance_adj = dispatch["net_receivable"] - slip.slip_amount
    await db.dairy_plants.update_one(
        {"id": dispatch["dairy_plant_id"]},
        {"$inc": {"total_amount": -new_balance_adj, "balance": -new_balance_adj}}
    )

    await db.dispatches.update_one(
        {"id": dispatch_id},
        {"$set": {
            "slip_fat": slip.slip_fat,
            "slip_snf": slip.slip_snf,
            "slip_amount": slip.slip_amount,
            "slip_deductions": slip.slip_deductions,
            "fat_difference": fat_diff,
            "amount_difference": amount_diff,
            "slip_matched": True,
        }}
    )

    updated = await db.dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    return DispatchResponse(**updated)

# ==================== DAIRY PAYMENT ROUTES ====================

@api_router.post("/dairy-payments", response_model=DairyPaymentResponse)
async def create_dairy_payment(payment: DairyPaymentCreate, current_user: dict = Depends(get_current_user)):
    plant = await db.dairy_plants.find_one({"id": payment.dairy_plant_id}, {"_id": 0})
    if not plant:
        raise HTTPException(status_code=404, detail="Dairy plant not found")

    payment_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc)
    payment_doc = {
        "id": payment_id,
        "dairy_plant_id": payment.dairy_plant_id,
        "dairy_plant_name": plant["name"],
        "amount": payment.amount,
        "payment_mode": payment.payment_mode,
        "reference_number": payment.reference_number or "",
        "notes": payment.notes or "",
        "date": now.strftime("%Y-%m-%d"),
        "created_at": now.isoformat(),
    }
    await db.dairy_payments.insert_one(payment_doc)
    del payment_doc["_id"]

    await db.dairy_plants.update_one(
        {"id": payment.dairy_plant_id},
        {"$inc": {"total_paid": payment.amount, "balance": -payment.amount}}
    )

    return DairyPaymentResponse(**payment_doc)

@api_router.get("/dairy-payments")
async def get_dairy_payments(dairy_plant_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if dairy_plant_id:
        query["dairy_plant_id"] = dairy_plant_id
    payments = await db.dairy_payments.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    return payments

# ==================== DAIRY LEDGER & PROFIT ROUTES ====================

@api_router.get("/dairy-plants/{plant_id}/ledger")
async def get_dairy_ledger(plant_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    plant = await db.dairy_plants.find_one({"id": plant_id}, {"_id": 0})
    if not plant:
        raise HTTPException(status_code=404, detail="Dairy plant not found")

    dq = {"dairy_plant_id": plant_id}
    if start_date:
        dq.setdefault("date", {})["$gte"] = start_date
    if end_date:
        dq.setdefault("date", {})["$lte"] = end_date

    dispatches = await db.dispatches.find(dq, {"_id": 0}).sort("date", -1).to_list(500)
    payments = await db.dairy_payments.find(dq, {"_id": 0}).sort("date", -1).to_list(500)

    return {"plant": DairyPlantResponse(**plant).model_dump(), "dispatches": dispatches, "payments": payments}

@api_router.get("/dairy/profit-report")
async def dairy_profit_report(date: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Calculate real net profit: Dispatch income - Farmer costs - Expenses"""
    if date:
        start_date = date
        end_date = date

    if not start_date:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not end_date:
        end_date = start_date

    # Get dispatch income
    dispatch_query = {"date": {"$gte": start_date, "$lte": end_date}}
    dispatches = await db.dispatches.find(dispatch_query, {"_id": 0}).to_list(1000)
    total_dispatch_amount = sum(d.get("net_receivable", 0) for d in dispatches)
    total_dispatch_kg = sum(d.get("quantity_kg", 0) for d in dispatches)
    avg_selling_rate = round(total_dispatch_amount / total_dispatch_kg, 2) if total_dispatch_kg > 0 else 0

    # Get farmer purchase cost
    collection_query = {"date": {"$gte": start_date, "$lte": end_date}}
    collections = await db.milk_collections.find(collection_query, {"_id": 0}).to_list(5000)
    total_farmer_amount = sum(c.get("amount", 0) for c in collections)
    total_collection_liters = sum(c.get("quantity", 0) for c in collections)
    total_collection_kg = round(total_collection_liters * 1.03, 2)  # Convert liters to KG
    avg_buying_rate = round(total_farmer_amount / total_collection_liters, 2) if total_collection_liters > 0 else 0

    # Weighted FAT average from collections
    weighted_fat = sum(c.get("fat", 0) * c.get("quantity", 0) for c in collections)
    avg_collection_fat = round(weighted_fat / total_collection_liters, 2) if total_collection_liters > 0 else 0
    weighted_snf = sum(c.get("snf", 0) * c.get("quantity", 0) for c in collections)
    avg_collection_snf = round(weighted_snf / total_collection_liters, 2) if total_collection_liters > 0 else 0

    # Dispatch FAT average
    dispatch_weighted_fat = sum(d.get("avg_fat", 0) * d.get("quantity_kg", 0) for d in dispatches)
    avg_dispatch_fat = round(dispatch_weighted_fat / total_dispatch_kg, 2) if total_dispatch_kg > 0 else 0

    # Milk loss/gain
    milk_difference_kg = round(total_collection_kg - total_dispatch_kg, 2)
    milk_loss_percent = round((milk_difference_kg / total_collection_kg) * 100, 2) if total_collection_kg > 0 else 0

    # Get expenses
    expense_query = {"date": {"$gte": start_date, "$lte": end_date}}
    expenses = await db.expenses.find(expense_query, {"_id": 0}).to_list(1000)
    total_expenses = sum(e.get("amount", 0) for e in expenses)
    expense_by_category = {}
    for e in expenses:
        cat = e.get("category", "other")
        expense_by_category[cat] = expense_by_category.get(cat, 0) + e.get("amount", 0)

    # Dispatch deductions breakdown
    total_dispatch_deductions = sum(d.get("total_deduction", 0) for d in dispatches)

    # Get retail/shop sales
    sales_query = {"date": {"$gte": start_date, "$lte": end_date}}
    sales = await db.sales.find(sales_query, {"_id": 0}).to_list(5000)
    total_retail_sales = sum(s.get("amount", 0) for s in sales)
    retail_milk_sales = sum(s.get("amount", 0) for s in sales if s.get("product") == "milk")
    retail_other_sales = total_retail_sales - retail_milk_sales

    # Calculate profits (include retail sales)
    total_income = total_dispatch_amount + total_retail_sales
    gross_margin_per_unit = round(avg_selling_rate - avg_buying_rate, 2)
    gross_profit = round(total_income - total_farmer_amount, 2)
    net_profit = round(gross_profit - total_expenses, 2)

    return {
        "period": {"start_date": start_date, "end_date": end_date},
        "dispatch": {
            "total_kg": total_dispatch_kg,
            "total_amount": total_dispatch_amount,
            "avg_rate": avg_selling_rate,
            "avg_fat": avg_dispatch_fat,
            "total_deductions": total_dispatch_deductions,
            "count": len(dispatches),
        },
        "collection": {
            "total_liters": total_collection_liters,
            "total_kg": total_collection_kg,
            "total_amount": total_farmer_amount,
            "avg_rate": avg_buying_rate,
            "avg_fat": avg_collection_fat,
            "avg_snf": avg_collection_snf,
            "count": len(collections),
        },
        "retail_sales": {
            "total_amount": total_retail_sales,
            "milk_sales": retail_milk_sales,
            "other_sales": retail_other_sales,
            "count": len(sales),
        },
        "milk_tracking": {
            "collected_kg": total_collection_kg,
            "dispatched_kg": total_dispatch_kg,
            "difference_kg": milk_difference_kg,
            "loss_percent": milk_loss_percent,
            "alert": milk_loss_percent > 1,
        },
        "fat_analysis": {
            "collection_avg_fat": avg_collection_fat,
            "dispatch_avg_fat": avg_dispatch_fat,
            "fat_deviation": round(avg_collection_fat - avg_dispatch_fat, 2),
        },
        "expenses": {
            "total": total_expenses,
            "by_category": expense_by_category,
        },
        "profit": {
            "total_income": total_income,
            "gross_margin_per_unit": gross_margin_per_unit,
            "gross_profit": gross_profit,
            "net_profit": net_profit,
        },
    }

@api_router.get("/dairy/fat-analysis")
async def fat_analysis_report(start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Farmer-wise FAT analysis for quality tracking"""
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    collections = await db.milk_collections.find(
        {"date": {"$gte": start_date, "$lte": end_date}}, {"_id": 0}
    ).to_list(10000)

    farmer_stats = {}
    for c in collections:
        fid = c["farmer_id"]
        if fid not in farmer_stats:
            farmer_stats[fid] = {"farmer_id": fid, "farmer_name": c["farmer_name"], "total_qty": 0, "weighted_fat": 0, "entries": 0, "milk_type": c.get("milk_type", "cow")}
        farmer_stats[fid]["total_qty"] += c["quantity"]
        farmer_stats[fid]["weighted_fat"] += c["fat"] * c["quantity"]
        farmer_stats[fid]["entries"] += 1

    result = []
    for fid, stats in farmer_stats.items():
        avg_fat = round(stats["weighted_fat"] / stats["total_qty"], 2) if stats["total_qty"] > 0 else 0
        result.append({
            "farmer_id": stats["farmer_id"],
            "farmer_name": stats["farmer_name"],
            "milk_type": stats["milk_type"],
            "total_quantity": round(stats["total_qty"], 1),
            "avg_fat": avg_fat,
            "entries": stats["entries"],
            "quality": "good" if avg_fat >= 4.0 else ("average" if avg_fat >= 3.0 else "low"),
        })

    result.sort(key=lambda x: x["avg_fat"], reverse=True)

    overall_fat = round(sum(c["fat"] * c["quantity"] for c in collections) / sum(c["quantity"] for c in collections), 2) if collections else 0

    return {
        "period": {"start_date": start_date, "end_date": end_date},
        "overall_avg_fat": overall_fat,
        "total_farmers": len(result),
        "quality_breakdown": {
            "good": len([r for r in result if r["quality"] == "good"]),
            "average": len([r for r in result if r["quality"] == "average"]),
            "low": len([r for r in result if r["quality"] == "low"]),
        },
        "farmers": result,
    }

# ==================== DISPATCH BILL / PRINT ROUTES ====================

@api_router.get("/dispatches/{dispatch_id}/bill")
async def get_dispatch_bill(dispatch_id: str, current_user: dict = Depends(get_current_user)):
    dispatch = await db.dispatches.find_one({"id": dispatch_id}, {"_id": 0})
    if not dispatch:
        raise HTTPException(status_code=404, detail="Dispatch not found")
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    dairy_phone = settings.get("phone", "")
    dairy_address = settings.get("address", "")
    d = dispatch
    ded_rows = ""
    for ded in d.get("deductions", []):
        ded_rows += f"<tr><td>{ded['type'].replace('_',' ').title()}</td><td class='right'>-₹{ded['amount']:.2f}</td></tr>"
    slip_section = ""
    if d.get("slip_matched"):
        slip_section = f"""<h3>Dairy Slip Comparison / डेयरी स्लिप तुलना</h3>
        <table><tr><th></th><th>Your / आपका</th><th>Dairy Slip / डेयरी स्लिप</th><th>Diff / अंतर</th></tr>
        <tr><td>FAT %</td><td>{d['avg_fat']}</td><td>{d.get('slip_fat','N/A')}</td><td>{d.get('fat_difference',0)}</td></tr>
        <tr><td>Amount / राशि</td><td>₹{d['net_receivable']:.2f}</td><td>₹{d.get('slip_amount',0):.2f}</td><td>₹{d.get('amount_difference',0):.2f}</td></tr></table>"""
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>
*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1a1a1a;padding:15mm 20mm;max-width:210mm}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #15803d;padding-bottom:12px;margin-bottom:15px}}
.dairy-name{{font-size:24px;font-weight:bold;color:#15803d}}.dairy-info{{font-size:10px;color:#666;margin-top:4px}}
.invoice-box{{text-align:right}}.invoice-title{{font-size:20px;color:#15803d;font-weight:bold}}
.info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:15px;margin-bottom:15px}}
.info-card{{background:#f8faf8;border:1px solid #e5e7eb;border-radius:6px;padding:10px}}
.info-label{{font-size:9px;color:#888;text-transform:uppercase}}.info-value{{font-size:13px;font-weight:600;margin-top:2px}}
table{{width:100%;border-collapse:collapse;margin-bottom:15px}}
th{{background:#15803d;color:white;padding:6px 8px;text-align:left;font-size:10px}}td{{padding:5px 8px;border-bottom:1px solid #eee;font-size:11px}}.right{{text-align:right}}
.summary-grid{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;margin:15px 0}}
.summary-box{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:10px;text-align:center}}
.summary-label{{font-size:9px;color:#666;text-transform:uppercase}}.summary-value{{font-size:18px;font-weight:bold;color:#15803d;margin-top:2px}}
h3{{color:#15803d;font-size:13px;margin:12px 0 8px;padding-bottom:4px;border-bottom:1px solid #d1fae5}}
.sig-box{{border-top:1px solid #333;width:150px;text-align:center;padding-top:5px;margin-top:40px;font-size:10px}}
@media print{{@page{{size:A4;margin:10mm}}body{{padding:0}}}}
</style></head><body>
<div class="header"><div><div class="dairy-name">{dairy_name}</div><div class="dairy-info">{dairy_address}<br>{dairy_phone}</div></div>
<div class="invoice-box"><div class="invoice-title">DISPATCH BILL / डिस्पैच बिल</div><div style="font-size:11px;color:#666;margin-top:2px">{d['date']}</div></div></div>
<div class="info-grid">
<div class="info-card"><div class="info-label">Dairy Plant / डेयरी प्लांट</div><div class="info-value">{d['dairy_plant_name']}</div></div>
<div class="info-card"><div class="info-label">Tanker / टैंकर</div><div class="info-value">{d.get('tanker_number','N/A')}</div></div>
</div>
<div class="summary-grid">
<div class="summary-box"><div class="summary-label">Quantity / मात्रा</div><div class="summary-value">{d['quantity_kg']} KG</div></div>
<div class="summary-box"><div class="summary-label">FAT %</div><div class="summary-value">{d['avg_fat']}%</div></div>
<div class="summary-box"><div class="summary-label">SNF %</div><div class="summary-value">{d['avg_snf']}%</div></div>
<div class="summary-box"><div class="summary-label">Rate / दर</div><div class="summary-value">₹{d['rate_per_kg']}/KG</div></div>
</div>
<h3>Amount Calculation / राशि गणना</h3>
<table><tr><td>Gross Amount / कुल राशि</td><td class="right" style="font-weight:bold">₹{d['gross_amount']:.2f}</td></tr>
{ded_rows}
<tr style="background:#15803d;color:white;font-weight:bold"><td>Net Receivable / शुद्ध प्राप्य</td><td class="right">₹{d['net_receivable']:.2f}</td></tr></table>
{slip_section}
<div style="display:flex;justify-content:flex-end;gap:40px;margin-top:30px">
<div class="sig-box">Dairy Stamp / डेयरी मुहर</div><div class="sig-box">Signature / हस्ताक्षर</div></div>
</body></html>"""
    return {"html": html, "dispatch": dispatch}

@api_router.get("/dairy-plants/{plant_id}/statement")
async def get_dairy_statement(plant_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    plant = await db.dairy_plants.find_one({"id": plant_id}, {"_id": 0})
    if not plant:
        raise HTTPException(status_code=404, detail="Dairy plant not found")
    if not start_date:
        start_date = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    dq = {"dairy_plant_id": plant_id, "date": {"$gte": start_date, "$lte": end_date}}
    dispatches = await db.dispatches.find(dq, {"_id": 0}).sort("date", 1).to_list(500)
    payments = await db.dairy_payments.find(dq, {"_id": 0}).sort("date", 1).to_list(500)
    settings = await db.settings.find_one({"type": "dairy_info"}, {"_id": 0}) or {}
    dairy_name = settings.get("dairy_name", "Nirbani Dairy")
    dairy_phone = settings.get("phone", "")
    dairy_address = settings.get("address", "")
    total_supplied = sum(d["quantity_kg"] for d in dispatches)
    total_amount = sum(d["net_receivable"] for d in dispatches)
    total_paid = sum(p["amount"] for p in payments)
    balance = total_amount - total_paid
    disp_rows = ""
    for i, d in enumerate(dispatches, 1):
        disp_rows += f"<tr><td>{i}</td><td>{d['date']}</td><td>{d.get('tanker_number','')}</td><td>{d['quantity_kg']}</td><td>{d['avg_fat']}%</td><td>₹{d['rate_per_kg']}</td><td class='right'>-₹{d['total_deduction']:.0f}</td><td class='right'>₹{d['net_receivable']:.2f}</td></tr>"
    pay_rows = ""
    for p in payments:
        pay_rows += f"<tr><td>{p['date']}</td><td>{p['payment_mode'].upper()}</td><td>{p.get('reference_number','')}</td><td class='right'>₹{p['amount']:.2f}</td></tr>"
    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"><style>
*{{margin:0;padding:0;box-sizing:border-box}}body{{font-family:'Segoe UI',Arial,sans-serif;font-size:12px;color:#1a1a1a;padding:15mm 20mm;max-width:210mm}}
.header{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:3px solid #15803d;padding-bottom:12px;margin-bottom:15px}}
.dairy-name{{font-size:24px;font-weight:bold;color:#15803d}}.dairy-info{{font-size:10px;color:#666;margin-top:4px}}
.invoice-box{{text-align:right}}.invoice-title{{font-size:20px;color:#15803d;font-weight:bold}}
.info-card{{background:#f8faf8;border:1px solid #e5e7eb;border-radius:6px;padding:10px;margin-bottom:15px}}
.info-label{{font-size:9px;color:#888;text-transform:uppercase}}.info-value{{font-size:13px;font-weight:600;margin-top:2px}}
table{{width:100%;border-collapse:collapse;margin-bottom:15px}}
th{{background:#15803d;color:white;padding:6px 8px;text-align:left;font-size:10px}}td{{padding:5px 8px;border-bottom:1px solid #eee;font-size:11px}}.right{{text-align:right}}
.summary-grid{{display:grid;grid-template-columns:1fr 1fr 1fr 1fr;gap:10px;margin:15px 0}}
.summary-box{{background:#f0fdf4;border:1px solid #bbf7d0;border-radius:6px;padding:10px;text-align:center}}
.summary-box.due{{background:#fef2f2;border-color:#fecaca}}
.summary-label{{font-size:9px;color:#666;text-transform:uppercase}}.summary-value{{font-size:18px;font-weight:bold;color:#15803d;margin-top:2px}}
.summary-box.due .summary-value{{color:#dc2626}}
h3{{color:#15803d;font-size:13px;margin:12px 0 8px;padding-bottom:4px;border-bottom:1px solid #d1fae5}}
.sig-box{{border-top:1px solid #333;width:150px;text-align:center;padding-top:5px;margin-top:40px;font-size:10px}}
@media print{{@page{{size:A4;margin:10mm}}body{{padding:0}}}}
</style></head><body>
<div class="header"><div><div class="dairy-name">{dairy_name}</div><div class="dairy-info">{dairy_address}<br>{dairy_phone}</div></div>
<div class="invoice-box"><div class="invoice-title">DAIRY STATEMENT / डेयरी विवरण</div><div style="font-size:11px;color:#666;margin-top:2px">{start_date} to {end_date}</div></div></div>
<div class="info-card"><div class="info-label">Dairy Plant / डेयरी प्लांट</div><div class="info-value">{plant['name']}</div>
<div style="font-size:10px;color:#666">{plant.get('address','')}{f" | Ph: {plant.get('phone','')}" if plant.get('phone') else ''}</div></div>
<div class="summary-grid">
<div class="summary-box"><div class="summary-label">Total Supplied / कुल आपूर्ति</div><div class="summary-value">{total_supplied:.1f} KG</div></div>
<div class="summary-box"><div class="summary-label">Total Amount / कुल राशि</div><div class="summary-value">₹{total_amount:.0f}</div></div>
<div class="summary-box"><div class="summary-label">Total Paid / कुल भुगतान</div><div class="summary-value">₹{total_paid:.0f}</div></div>
<div class="summary-box {'due' if balance > 0 else ''}"><div class="summary-label">Balance / बकाया</div><div class="summary-value">₹{balance:.0f}</div></div>
</div>
<h3>Dispatch Details / डिस्पैच विवरण</h3>
<table><tr><th>#</th><th>Date</th><th>Tanker</th><th>Qty KG</th><th>FAT</th><th>Rate</th><th class="right">Deductions</th><th class="right">Net ₹</th></tr>
{disp_rows}
<tr style="background:#15803d;color:white;font-weight:bold"><td colspan="3">TOTAL</td><td>{total_supplied:.1f}</td><td></td><td></td><td></td><td class="right">₹{total_amount:.2f}</td></tr></table>
{f'<h3>Payments / भुगतान</h3><table><tr><th>Date</th><th>Mode</th><th>Reference</th><th class="right">Amount ₹</th></tr>{pay_rows}<tr style="background:#15803d;color:white;font-weight:bold"><td colspan="3">TOTAL PAID</td><td class="right">₹{total_paid:.2f}</td></tr></table>' if payments else ''}
<div style="display:flex;justify-content:flex-end;gap:40px;margin-top:30px">
<div class="sig-box">Dairy Stamp / डेयरी मुहर</div><div class="sig-box">Signature / हस्ताक्षर</div></div>
</body></html>"""
    return {"html": html}

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "Nirbani Dairy Management System API", "status": "healthy"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def seed_admin_user():
    admin_email = "nirbanidairy@gmal.com"
    existing = await db.users.find_one({"email": admin_email}, {"_id": 0})
    if not existing:
        admin_doc = {
            "id": str(uuid.uuid4()),
            "name": "Nirbani Admin",
            "email": admin_email,
            "phone": "0000000000",
            "password": hash_password("Nirbani0056!"),
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "is_active": True
        }
        await db.users.insert_one(admin_doc)
        logger.info("Admin user seeded successfully")
    else:
        if existing.get("role") != "admin":
            await db.users.update_one({"email": admin_email}, {"$set": {"role": "admin"}})
            logger.info("Admin role updated for existing user")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
